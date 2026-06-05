import os
import io
import smtplib
import warnings
from datetime import datetime, timedelta
from dataclasses import dataclass, field, asdict
from typing import List, Dict, Optional, Any, Tuple
from collections import defaultdict
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib
from matplotlib import font_manager
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, mm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    PageBreak, Image, KeepTogether, ListFlowable, ListItem
)
from reportlab.pdfgen import canvas
from reportlab.graphics.shapes import Drawing, Rect, String
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from sqlalchemy.orm import Session
from sqlalchemy import func, and_, or_

from config import config
from models import (
    Vulnerability, VulnerabilityInstance, WorkOrder, Asset,
    SeverityEnum, WorkOrderStatusEnum, Report, ReportTypeEnum,
    ReviewTask, VerificationRecord
)
from database import db_manager, with_read_session, with_session
from logger import logger, log_with_context

warnings.filterwarnings("ignore")

matplotlib.use("Agg")

SEVERITY_COLORS = {
    "critical": "#DC2626",
    "high": "#EA580C",
    "medium": "#CA8A04",
    "low": "#16A34A"
}

SEVERITY_NAMES = {
    "critical": "严重",
    "high": "高危",
    "medium": "中危",
    "low": "低危"
}

STAGE_NAMES = {
    "pending": "待处理",
    "fixing": "修复中",
    "fixed": "已修复",
    "verifying": "验证中",
    "closed": "已关闭"
}


def _setup_chinese_font():
    font_paths = [
        "/System/Library/Fonts/PingFang.ttc",
        "/System/Library/Fonts/STHeiti Medium.ttc",
        "/Library/Fonts/Arial Unicode.ttf",
        "/usr/share/fonts/truetype/wqy/wqy-microhei.ttc",
        "C:/Windows/Fonts/msyh.ttc",
        "C:/Windows/Fonts/simhei.ttf"
    ]
    for font_path in font_paths:
        if os.path.exists(font_path):
            try:
                font_manager.fontManager.addfont(font_path)
                font_name = font_manager.FontProperties(fname=font_path).get_name()
                matplotlib.rcParams["font.sans-serif"] = [font_name]
                matplotlib.rcParams["axes.unicode_minus"] = False
                return font_name
            except Exception:
                continue
    matplotlib.rcParams["font.sans-serif"] = ["DejaVu Sans"]
    matplotlib.rcParams["axes.unicode_minus"] = False
    return "DejaVu Sans"


CHINESE_FONT = _setup_chinese_font()


@dataclass
class DailyStatsData:
    date: str
    total_vulns: int = 0
    new_vulns: int = 0
    repeated_vulns: int = 0
    high_risk_count: int = 0
    vulns_by_severity: Dict[str, int] = field(default_factory=dict)
    vulns_by_asset_type: Dict[str, int] = field(default_factory=dict)
    vulns_by_department: Dict[str, int] = field(default_factory=dict)
    vulns_by_vuln_type: Dict[str, int] = field(default_factory=dict)
    total_work_orders: int = 0
    pending_count: int = 0
    fixing_count: int = 0
    fixed_count: int = 0
    verifying_count: int = 0
    closed_count: int = 0
    fix_rate: float = 0.0
    avg_fix_duration_hours: float = 0.0
    timeout_rate: float = 0.0
    timeout_count: int = 0
    avg_stage_durations: Dict[str, float] = field(default_factory=dict)
    verify_fail_count: int = 0
    verify_fail_rate: float = 0.0
    review_tasks_count: int = 0
    stage_distribution: Dict[str, int] = field(default_factory=dict)
    vuln_asset_matrix: Dict[str, Dict[str, int]] = field(default_factory=dict)

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


@dataclass
class TrendData:
    period: str
    granularity: str
    periods: List[str] = field(default_factory=list)
    values: List[float] = field(default_factory=list)
    values_last_period: List[float] = field(default_factory=list)
    mom_change: float = 0.0
    yoy_change: float = 0.0
    forecast: List[float] = field(default_factory=list)
    anomalies: List[Tuple[str, float, str]] = field(default_factory=list)
    metric_name: str = ""
    metric_unit: str = ""

    def to_dict(self) -> Dict[str, Any]:
        return asdict(self)


class StatsEngine:
    def __init__(self):
        self.stages = ["pending", "fixing", "fixed", "verifying", "closed"]

    @with_read_session
    def calculate_daily_stats(self, date: datetime, session: Optional[Session] = None) -> DailyStatsData:
        if session is None:
            raise ValueError("Session is required")

        date_start = datetime(date.year, date.month, date.day)
        date_end = date_start + timedelta(days=1)

        stats = DailyStatsData(date=date_start.strftime("%Y-%m-%d"))

        vuln_instances = session.query(VulnerabilityInstance).filter(
            and_(
                VulnerabilityInstance.discovery_time >= date_start,
                VulnerabilityInstance.discovery_time < date_end
            )
        ).all()

        stats.total_vulns = len(vuln_instances)

        vuln_ids = [vi.vuln_id for vi in vuln_instances]
        earlier_instances = session.query(VulnerabilityInstance.vuln_id).filter(
            and_(
                VulnerabilityInstance.vuln_id.in_(vuln_ids),
                VulnerabilityInstance.discovery_time < date_start
            )
        ).distinct().all()
        repeated_ids = {row[0] for row in earlier_instances}

        stats.new_vulns = len([vi for vi in vuln_instances if vi.vuln_id not in repeated_ids])
        stats.repeated_vulns = stats.total_vulns - stats.new_vulns

        for vi in vuln_instances:
            severity = vi.vulnerability.severity.value if vi.vulnerability else "unknown"
            stats.vulns_by_severity[severity] = stats.vulns_by_severity.get(severity, 0) + 1

            if severity in ["critical", "high"]:
                stats.high_risk_count += 1

            asset_type = vi.asset.type if vi.asset else "unknown"
            stats.vulns_by_asset_type[asset_type] = stats.vulns_by_asset_type.get(asset_type, 0) + 1

            department = vi.asset.department if vi.asset else "unknown"
            stats.vulns_by_department[department] = stats.vulns_by_department.get(department, 0) + 1

            vuln_type = vi.vulnerability.extra_data.get("vuln_type", "other") if vi.vulnerability and vi.vulnerability.extra_data else "other"
            stats.vulns_by_vuln_type[vuln_type] = stats.vulns_by_vuln_type.get(vuln_type, 0) + 1

            if vuln_type not in stats.vuln_asset_matrix:
                stats.vuln_asset_matrix[vuln_type] = {}
            stats.vuln_asset_matrix[vuln_type][asset_type] = stats.vuln_asset_matrix[vuln_type].get(asset_type, 0) + 1

        work_orders = session.query(WorkOrder).filter(
            and_(
                WorkOrder.created_at >= date_start,
                WorkOrder.created_at < date_end
            )
        ).all()

        stats.total_work_orders = len(work_orders)

        for wo in work_orders:
            status = wo.status.value
            stats.stage_distribution[status] = stats.stage_distribution.get(status, 0) + 1

            if status == "pending":
                stats.pending_count += 1
            elif status == "fixing":
                stats.fixing_count += 1
            elif status == "fixed":
                stats.fixed_count += 1
            elif status == "verifying":
                stats.verifying_count += 1
            elif status == "closed":
                stats.closed_count += 1

            if wo.closed_at and wo.created_at:
                if wo.closed_at > wo.deadline:
                    stats.timeout_count += 1

        closed_orders = session.query(WorkOrder).filter(
            and_(
                WorkOrder.closed_at >= date_start,
                WorkOrder.closed_at < date_end,
                WorkOrder.status == WorkOrderStatusEnum.CLOSED
            )
        ).all()

        if closed_orders:
            total_fix_duration = 0
            valid_count = 0
            for wo in closed_orders:
                if wo.closed_at and wo.created_at:
                    duration = (wo.closed_at - wo.created_at).total_seconds() / 3600
                    total_fix_duration += duration
                    valid_count += 1
            if valid_count > 0:
                stats.avg_fix_duration_hours = total_fix_duration / valid_count

            fix_count = len([wo for wo in closed_orders if wo.fixed_at])
            stats.fix_rate = fix_count / len(closed_orders) if closed_orders else 0.0

        if stats.total_work_orders > 0:
            stats.timeout_rate = stats.timeout_count / stats.total_work_orders

        stats.avg_stage_durations = self._calculate_stage_durations(date_start, date_end, session)

        verify_records = session.query(VerificationRecord).filter(
            and_(
                VerificationRecord.verification_time >= date_start,
                VerificationRecord.verification_time < date_end
            )
        ).all()
        stats.verify_fail_count = len([vr for vr in verify_records if not vr.is_fixed])
        if verify_records:
            stats.verify_fail_rate = stats.verify_fail_count / len(verify_records)

        stats.review_tasks_count = session.query(ReviewTask).filter(
            and_(
                ReviewTask.created_at >= date_start,
                ReviewTask.created_at < date_end
            )
        ).count()

        return stats

    def _calculate_stage_durations(self, date_start: datetime, date_end: datetime, session: Session) -> Dict[str, float]:
        stage_durations = {}
        for stage in self.stages:
            total_duration = 0
            count = 0

            orders = session.query(WorkOrder).filter(
                WorkOrder.closed_at >= date_start,
                WorkOrder.closed_at < date_end
            ).all()

            for wo in orders:
                duration = self._get_stage_duration(wo, stage)
                if duration is not None:
                    total_duration += duration
                    count += 1

            if count > 0:
                stage_durations[stage] = total_duration / count / 3600
            else:
                stage_durations[stage] = 0.0

        return stage_durations

    def _get_stage_duration(self, work_order: WorkOrder, stage: str) -> Optional[float]:
        timestamps = {
            "pending": work_order.created_at,
            "fixing": work_order.started_at,
            "fixed": work_order.fixed_at,
            "verifying": work_order.verified_at,
            "closed": work_order.closed_at
        }

        current_time = timestamps.get(stage)
        if current_time is None:
            return None

        stage_idx = self.stages.index(stage)
        if stage_idx < len(self.stages) - 1:
            next_stage = self.stages[stage_idx + 1]
            next_time = timestamps.get(next_stage)
            if next_time:
                return (next_time - current_time).total_seconds()

        return None

    @with_read_session
    def calculate_custom_stats(self, start_date: datetime, end_date: datetime,
                                group_by: Optional[List[str]] = None,
                                session: Optional[Session] = None) -> Dict[str, Any]:
        if session is None:
            raise ValueError("Session is required")

        group_by = group_by or ["department", "asset_type", "severity"]
        results = {}

        vuln_query = session.query(VulnerabilityInstance).join(Asset).join(Vulnerability).filter(
            and_(
                VulnerabilityInstance.discovery_time >= start_date,
                VulnerabilityInstance.discovery_time < end_date
            )
        )

        vuln_instances = vuln_query.all()
        results["total_vulns"] = len(vuln_instances)

        for group in group_by:
            group_data = defaultdict(int)
            for vi in vuln_instances:
                if group == "department":
                    key = vi.asset.department if vi.asset else "unknown"
                elif group == "asset_type":
                    key = vi.asset.type if vi.asset else "unknown"
                elif group == "severity":
                    key = vi.vulnerability.severity.value if vi.vulnerability else "unknown"
                elif group == "vuln_type":
                    key = vi.vulnerability.extra_data.get("vuln_type", "other") if vi.vulnerability and vi.vulnerability.extra_data else "other"
                else:
                    key = "unknown"
                group_data[key] += 1
            results[f"by_{group}"] = dict(group_data)

        wo_query = session.query(WorkOrder).filter(
            and_(
                WorkOrder.created_at >= start_date,
                WorkOrder.created_at < end_date
            )
        )
        work_orders = wo_query.all()
        results["total_work_orders"] = len(work_orders)
        results["closed_count"] = len([wo for wo in work_orders if wo.status == WorkOrderStatusEnum.CLOSED])
        results["fix_rate"] = results["closed_count"] / len(work_orders) if work_orders else 0.0

        return results


class TrendAnalyzer:
    def __init__(self, anomaly_threshold: float = 2.0):
        self.anomaly_threshold = anomaly_threshold
        self.stats_engine = StatsEngine()

    @with_read_session
    def generate_trend_data(self, start_date: datetime, end_date: datetime,
                             granularity: str = "day", metric: str = "new_vulns",
                             session: Optional[Session] = None) -> TrendData:
        if session is None:
            raise ValueError("Session is required")

        granularity = granularity.lower()
        periods, period_dates = self._generate_periods(start_date, end_date, granularity)

        values = []
        for period_start, period_end in period_dates:
            stats = self.stats_engine.calculate_custom_stats(period_start, period_end, session=session)
            values.append(self._extract_metric(stats, metric))

        last_period_start, last_period_end = self._get_previous_period(start_date, end_date, granularity)
        last_periods, _ = self._generate_periods(last_period_start, last_period_end, granularity)
        values_last = []
        for ps, pe in self._generate_periods(last_period_start, last_period_end, granularity)[1]:
            stats = self.stats_engine.calculate_custom_stats(ps, pe, session=session)
            values_last.append(self._extract_metric(stats, metric))

        mom_change = self._calculate_mom_change(values, values_last)
        yoy_change = self._calculate_yoy_change(values, values_last)

        forecast = self._simple_moving_average_forecast(values, periods=7)

        anomalies = self._detect_anomalies(values, periods)

        trend = TrendData(
            period=f"{start_date.strftime('%Y-%m-%d')} ~ {end_date.strftime('%Y-%m-%d')}",
            granularity=granularity,
            periods=periods,
            values=values,
            values_last_period=values_last,
            mom_change=mom_change,
            yoy_change=yoy_change,
            forecast=forecast,
            anomalies=anomalies,
            metric_name=self._get_metric_name(metric),
            metric_unit=self._get_metric_unit(metric)
        )

        return trend

    def _generate_periods(self, start_date: datetime, end_date: datetime, granularity: str) -> Tuple[List[str], List[Tuple[datetime, datetime]]]:
        periods = []
        period_dates = []
        current = start_date

        while current < end_date:
            if granularity == "day":
                period_end = current + timedelta(days=1)
                period_label = current.strftime("%Y-%m-%d")
            elif granularity == "week":
                period_end = current + timedelta(weeks=1)
                period_label = f"{current.strftime('%Y-%m-%d')}~{(current + timedelta(days=6)).strftime('%Y-%m-%d')}"
            elif granularity == "month":
                if current.month == 12:
                    period_end = datetime(current.year + 1, 1, 1)
                else:
                    period_end = datetime(current.year, current.month + 1, 1)
                period_label = current.strftime("%Y-%m")
            else:
                period_end = current + timedelta(days=1)
                period_label = current.strftime("%Y-%m-%d")

            periods.append(period_label)
            period_dates.append((current, period_end))
            current = period_end

        return periods, period_dates

    def _get_previous_period(self, start_date: datetime, end_date: datetime, granularity: str) -> Tuple[datetime, datetime]:
        duration = end_date - start_date
        if granularity == "day":
            prev_start = start_date - duration
            prev_end = start_date
        elif granularity == "week":
            prev_start = start_date - timedelta(weeks=1)
            prev_end = start_date
        elif granularity == "month":
            prev_start = start_date - timedelta(days=30)
            prev_end = start_date
        else:
            prev_start = start_date - duration
            prev_end = start_date
        return prev_start, prev_end

    def _extract_metric(self, stats: Dict[str, Any], metric: str) -> float:
        metric_map = {
            "new_vulns": "total_vulns",
            "total_vulns": "total_vulns",
            "work_orders": "total_work_orders",
            "fix_rate": "fix_rate",
            "closed_count": "closed_count"
        }
        return float(stats.get(metric_map.get(metric, metric), 0))

    def _calculate_mom_change(self, current: List[float], previous: List[float]) -> float:
        if not current or not previous:
            return 0.0
        curr_total = sum(current)
        prev_total = sum(previous)
        if prev_total == 0:
            return 100.0 if curr_total > 0 else 0.0
        return ((curr_total - prev_total) / prev_total) * 100

    def _calculate_yoy_change(self, current: List[float], previous: List[float]) -> float:
        return self._calculate_mom_change(current, previous)

    def _simple_moving_average_forecast(self, values: List[float], periods: int = 7) -> List[float]:
        if len(values) < 3:
            return []

        forecast = []
        window = min(periods, len(values) // 2)

        for i in range(len(values), len(values) + periods):
            if i < window:
                ma = sum(values[:i + 1]) / (i + 1)
            else:
                ma = sum(values[i - window:i]) / window
            forecast.append(round(ma, 2))

        return forecast

    def _detect_anomalies(self, values: List[float], periods: List[str]) -> List[Tuple[str, float, str]]:
        if len(values) < 4:
            return []

        anomalies = []
        mean = np.mean(values)
        std = np.std(values)

        if std == 0:
            return []

        for i, (period, value) in enumerate(zip(periods, values)):
            z_score = abs(value - mean) / std
            if z_score > self.anomaly_threshold:
                if value > mean:
                    anomaly_type = "突增"
                else:
                    anomaly_type = "突降"
                anomalies.append((period, round(value, 2), anomaly_type))

        return anomalies

    def _get_metric_name(self, metric: str) -> str:
        names = {
            "new_vulns": "新增漏洞数",
            "total_vulns": "总漏洞数",
            "work_orders": "工单数",
            "fix_rate": "修复率",
            "closed_count": "已关闭工单数"
        }
        return names.get(metric, metric)

    def _get_metric_unit(self, metric: str) -> str:
        if metric == "fix_rate":
            return "%"
        return "个"


class ChartGenerator:
    def __init__(self, output_dir: Optional[str] = None, dpi: int = 300):
        self.output_dir = output_dir or config.REPORT_OUTPUT_DIR
        self.dpi = dpi
        self._ensure_output_dir()

    def _ensure_output_dir(self):
        os.makedirs(self.output_dir, exist_ok=True)
        charts_dir = os.path.join(self.output_dir, "charts")
        os.makedirs(charts_dir, exist_ok=True)

    def generate_charts(self, stats_data: DailyStatsData, trend_data: Optional[TrendData] = None,
                        output_dir: Optional[str] = None) -> Dict[str, str]:
        if output_dir:
            self.output_dir = output_dir
            self._ensure_output_dir()

        chart_paths = {}

        chart_paths["severity_pie"] = self._generate_severity_pie_chart(stats_data.vulns_by_severity)
        chart_paths["asset_type_pie"] = self._generate_asset_type_pie_chart(stats_data.vulns_by_asset_type)
        chart_paths["daily_bar"] = self._generate_daily_bar_chart(stats_data)
        chart_paths["department_bar"] = self._generate_department_bar_chart(stats_data.vulns_by_department)
        chart_paths["stage_stacked_bar"] = self._generate_stage_stacked_bar_chart(stats_data.stage_distribution)

        if trend_data:
            chart_paths["trend_line"] = self._generate_trend_line_chart(trend_data)

        chart_paths["fix_duration_trend"] = self._generate_fix_duration_trend(stats_data.avg_stage_durations)
        chart_paths["heatmap"] = self._generate_heatmap(stats_data.vuln_asset_matrix)

        return chart_paths

    def _generate_severity_pie_chart(self, data: Dict[str, int]) -> str:
        fig, ax = plt.subplots(figsize=(8, 8))

        labels = [SEVERITY_NAMES.get(k, k) for k in data.keys()]
        sizes = list(data.values())
        colors = [SEVERITY_COLORS.get(k, "#808080") for k in data.keys()]
        explode = [0.05] * len(data)

        if sum(sizes) > 0:
            wedges, texts, autotexts = ax.pie(
                sizes, explode=explode, labels=labels, colors=colors,
                autopct="%1.1f%%", shadow=True, startangle=90,
                textprops={"fontsize": 12}
            )
            for autotext in autotexts:
                autotext.set_color("white")
                autotext.set_fontweight("bold")
        else:
            ax.text(0.5, 0.5, "暂无数据", ha="center", va="center", fontsize=14)

        ax.set_title("漏洞严重等级分布", fontsize=16, fontweight="bold", pad=20)
        ax.axis("equal")

        filename = os.path.join(self.output_dir, "charts", "severity_distribution.png")
        plt.savefig(filename, dpi=self.dpi, bbox_inches="tight", facecolor="white")
        plt.close()
        return filename

    def _generate_asset_type_pie_chart(self, data: Dict[str, int]) -> str:
        fig, ax = plt.subplots(figsize=(8, 8))

        labels = list(data.keys())
        sizes = list(data.values())
        colors = plt.cm.Set3(np.linspace(0, 1, len(labels))) if labels else []

        if sum(sizes) > 0:
            wedges, texts, autotexts = ax.pie(
                sizes, labels=labels, colors=colors,
                autopct="%1.1f%%", shadow=True, startangle=90,
                textprops={"fontsize": 11}
            )
            for autotext in autotexts:
                autotext.set_color("white")
                autotext.set_fontweight("bold")
        else:
            ax.text(0.5, 0.5, "暂无数据", ha="center", va="center", fontsize=14)

        ax.set_title("资产类型分布", fontsize=16, fontweight="bold", pad=20)
        ax.axis("equal")
        ax.legend(loc="center left", bbox_to_anchor=(1, 0, 0.5, 1), fontsize=10)

        filename = os.path.join(self.output_dir, "charts", "asset_type_distribution.png")
        plt.savefig(filename, dpi=self.dpi, bbox_inches="tight", facecolor="white")
        plt.close()
        return filename

    def _generate_daily_bar_chart(self, stats: DailyStatsData) -> str:
        fig, ax = plt.subplots(figsize=(12, 6))

        categories = ["新增漏洞", "重复漏洞", "高危漏洞", "总工单"]
        values = [stats.new_vulns, stats.repeated_vulns, stats.high_risk_count, stats.total_work_orders]
        colors = ["#3B82F6", "#8B5CF6", "#DC2626", "#10B981"]

        bars = ax.bar(categories, values, color=colors, width=0.6)

        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2., height,
                    f"{int(height)}", ha="center", va="bottom", fontsize=12, fontweight="bold")

        ax.set_xlabel("类别", fontsize=12)
        ax.set_ylabel("数量", fontsize=12)
        ax.set_title(f"每日发现数量统计 ({stats.date})", fontsize=16, fontweight="bold", pad=20)
        ax.grid(axis="y", alpha=0.3)

        filename = os.path.join(self.output_dir, "charts", "daily_stats_bar.png")
        plt.savefig(filename, dpi=self.dpi, bbox_inches="tight", facecolor="white")
        plt.close()
        return filename

    def _generate_department_bar_chart(self, data: Dict[str, int]) -> str:
        fig, ax = plt.subplots(figsize=(12, 6))

        departments = list(data.keys())
        values = list(data.values())
        colors = plt.cm.viridis(np.linspace(0, 1, len(departments))) if departments else []

        bars = ax.barh(departments, values, color=colors)

        for bar in bars:
            width = bar.get_width()
            ax.text(width + 0.5, bar.get_y() + bar.get_height() / 2.,
                    f"{int(width)}", ha="left", va="center", fontsize=10)

        ax.set_xlabel("漏洞数量", fontsize=12)
        ax.set_ylabel("部门", fontsize=12)
        ax.set_title("各部门漏洞数量对比", fontsize=16, fontweight="bold", pad=20)
        ax.grid(axis="x", alpha=0.3)

        filename = os.path.join(self.output_dir, "charts", "department_comparison.png")
        plt.savefig(filename, dpi=self.dpi, bbox_inches="tight", facecolor="white")
        plt.close()
        return filename

    def _generate_stage_stacked_bar_chart(self, stage_data: Dict[str, int]) -> str:
        fig, ax = plt.subplots(figsize=(10, 6))

        stages = ["pending", "fixing", "fixed", "verifying", "closed"]
        stage_labels = [STAGE_NAMES.get(s, s) for s in stages]
        values = [stage_data.get(s, 0) for s in stages]
        colors = ["#6B7280", "#F59E0B", "#10B981", "#3B82F6", "#8B5CF6"]

        bottom = 0
        for i, (label, value, color) in enumerate(zip(stage_labels, values, colors)):
            if value > 0:
                ax.bar("工单阶段分布", value, bottom=bottom, label=label, color=color)
                ax.text("工单阶段分布", bottom + value / 2, f"{label}: {value}",
                        ha="center", va="center", fontsize=10, fontweight="bold", color="white")
            bottom += value

        ax.set_ylabel("工单数量", fontsize=12)
        ax.set_title("各阶段工单数量分布", fontsize=16, fontweight="bold", pad=20)
        ax.legend(loc="upper right", fontsize=10)
        ax.grid(axis="y", alpha=0.3)

        filename = os.path.join(self.output_dir, "charts", "stage_stacked_bar.png")
        plt.savefig(filename, dpi=self.dpi, bbox_inches="tight", facecolor="white")
        plt.close()
        return filename

    def _generate_trend_line_chart(self, trend: TrendData) -> str:
        fig, ax = plt.subplots(figsize=(14, 6))

        x = range(len(trend.periods))

        if trend.values:
            ax.plot(x, trend.values, marker="o", linewidth=2, label=f"本期{trend.metric_name}",
                    color="#3B82F6", markersize=6)

            for i, v in enumerate(trend.values):
                ax.annotate(f"{v:.0f}", (x[i], v), textcoords="offset points",
                            xytext=(0, 10), ha="center", fontsize=9)

        if trend.values_last_period and len(trend.values_last_period) == len(trend.values):
            ax.plot(x, trend.values_last_period, marker="s", linewidth=2, linestyle="--",
                    label=f"上期{trend.metric_name}", color="#9CA3AF", markersize=5, alpha=0.7)

        if trend.forecast:
            forecast_x = range(len(trend.values), len(trend.values) + len(trend.forecast))
            ax.plot(forecast_x, trend.forecast, marker="*", linewidth=2, linestyle=":",
                    label="预测值", color="#10B981", markersize=6)

        for period, value, atype in trend.anomalies:
            if period in trend.periods:
                idx = trend.periods.index(period)
                ax.annotate(f"{atype}!", (idx, value), textcoords="offset points",
                            xytext=(0, 15), ha="center", fontsize=10, color="red",
                            fontweight="bold",
                            bbox=dict(boxstyle="round,pad=0.3", fc="yellow", ec="red", alpha=0.8))

        ax.set_xlabel("时间周期", fontsize=12)
        ax.set_ylabel(f"{trend.metric_name} ({trend.metric_unit})", fontsize=12)
        ax.set_title(f"{trend.metric_name}趋势图", fontsize=16, fontweight="bold", pad=20)
        ax.set_xticks(list(x) + (list(forecast_x) if trend.forecast else []))
        all_labels = trend.periods + ([f"预测{i+1}" for i in range(len(trend.forecast))] if trend.forecast else [])
        ax.set_xticklabels(all_labels, rotation=45, ha="right", fontsize=9)
        ax.legend(loc="best", fontsize=10)
        ax.grid(True, alpha=0.3)

        if trend.mom_change != 0:
            mom_text = f"环比: {'+' if trend.mom_change > 0 else ''}{trend.mom_change:.1f}%"
            color = "red" if trend.mom_change > 0 else "green"
            ax.text(0.02, 0.95, mom_text, transform=ax.transAxes, fontsize=11,
                    color=color, fontweight="bold",
                    bbox=dict(boxstyle="round,pad=0.3", fc="white", ec=color, alpha=0.8))

        plt.tight_layout()
        filename = os.path.join(self.output_dir, "charts", "trend_line.png")
        plt.savefig(filename, dpi=self.dpi, bbox_inches="tight", facecolor="white")
        plt.close()
        return filename

    def _generate_fix_duration_trend(self, stage_durations: Dict[str, float]) -> str:
        fig, ax = plt.subplots(figsize=(10, 6))

        stages = ["pending", "fixing", "fixed", "verifying", "closed"]
        stage_labels = [STAGE_NAMES.get(s, s) for s in stages]
        durations = [stage_durations.get(s, 0) for s in stages]
        colors = ["#6B7280", "#F59E0B", "#10B981", "#3B82F6", "#8B5CF6"]

        bars = ax.bar(stage_labels, durations, color=colors, width=0.6)

        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width() / 2., height,
                    f"{height:.1f}h", ha="center", va="bottom", fontsize=11, fontweight="bold")

        ax.set_xlabel("修复阶段", fontsize=12)
        ax.set_ylabel("平均耗时 (小时)", fontsize=12)
        ax.set_title("各阶段平均修复时长", fontsize=16, fontweight="bold", pad=20)
        ax.grid(axis="y", alpha=0.3)

        filename = os.path.join(self.output_dir, "charts", "fix_duration_trend.png")
        plt.savefig(filename, dpi=self.dpi, bbox_inches="tight", facecolor="white")
        plt.close()
        return filename

    def _generate_heatmap(self, matrix: Dict[str, Dict[str, int]]) -> str:
        if not matrix:
            fig, ax = plt.subplots(figsize=(8, 6))
            ax.text(0.5, 0.5, "暂无数据", ha="center", va="center", fontsize=14)
            ax.axis("off")
            filename = os.path.join(self.output_dir, "charts", "heatmap.png")
            plt.savefig(filename, dpi=self.dpi, bbox_inches="tight", facecolor="white")
            plt.close()
            return filename

        vuln_types = list(matrix.keys())
        asset_types = sorted(list(set(at for v in matrix.values() for at in v.keys())))

        data_matrix = np.zeros((len(vuln_types), len(asset_types)))
        for i, vt in enumerate(vuln_types):
            for j, at in enumerate(asset_types):
                data_matrix[i, j] = matrix.get(vt, {}).get(at, 0)

        fig, ax = plt.subplots(figsize=(max(8, len(asset_types) * 1.5), max(6, len(vuln_types) * 1)))

        im = ax.imshow(data_matrix, cmap="YlOrRd", aspect="auto")

        ax.set_xticks(range(len(asset_types)))
        ax.set_yticks(range(len(vuln_types)))
        ax.set_xticklabels(asset_types, rotation=45, ha="right", fontsize=10)
        ax.set_yticklabels(vuln_types, fontsize=10)

        for i in range(len(vuln_types)):
            for j in range(len(asset_types)):
                value = int(data_matrix[i, j])
                if value > 0:
                    text_color = "white" if value > data_matrix.max() / 2 else "black"
                    ax.text(j, i, str(value), ha="center", va="center",
                            fontsize=9, color=text_color, fontweight="bold")

        cbar = ax.figure.colorbar(im, ax=ax)
        cbar.ax.set_ylabel("漏洞数量", rotation=-90, va="bottom", fontsize=10)

        ax.set_xlabel("资产类型", fontsize=12)
        ax.set_ylabel("漏洞类型", fontsize=12)
        ax.set_title("漏洞类型 × 资产类型 矩阵热力图", fontsize=14, fontweight="bold", pad=20)

        plt.tight_layout()
        filename = os.path.join(self.output_dir, "charts", "heatmap.png")
        plt.savefig(filename, dpi=self.dpi, bbox_inches="tight", facecolor="white")
        plt.close()
        return filename


class PdfReporter:
    def __init__(self, output_dir: Optional[str] = None):
        self.output_dir = output_dir or config.REPORT_OUTPUT_DIR
        self.styles = getSampleStyleSheet()
        self._setup_styles()
        self._ensure_output_dir()

    def _ensure_output_dir(self):
        os.makedirs(self.output_dir, exist_ok=True)

    def _setup_styles(self):
        self.title_style = ParagraphStyle(
            "CustomTitle",
            parent=self.styles["Heading1"],
            fontSize=24,
            textColor=colors.HexColor("#1E40AF"),
            alignment=1,
            spaceAfter=20,
            fontName="Helvetica-Bold"
        )

        self.heading2_style = ParagraphStyle(
            "CustomHeading2",
            parent=self.styles["Heading2"],
            fontSize=16,
            textColor=colors.HexColor("#1E3A8A"),
            spaceBefore=15,
            spaceAfter=10,
            fontName="Helvetica-Bold"
        )

        self.heading3_style = ParagraphStyle(
            "CustomHeading3",
            parent=self.styles["Heading3"],
            fontSize=14,
            textColor=colors.HexColor("#374151"),
            spaceBefore=10,
            spaceAfter=8,
            fontName="Helvetica-Bold"
        )

        self.body_style = ParagraphStyle(
            "CustomBody",
            parent=self.styles["BodyText"],
            fontSize=11,
            leading=16,
            textColor=colors.HexColor("#374151"),
            spaceAfter=8
        )

        self.highlight_style = ParagraphStyle(
            "Highlight",
            parent=self.body_style,
            textColor=colors.HexColor("#DC2626"),
            fontName="Helvetica-Bold"
        )

    def _header_footer(self, canvas_obj: canvas.Canvas, doc):
        canvas_obj.saveState()
        canvas_obj.setFont("Helvetica", 9)
        canvas_obj.setFillColor(colors.HexColor("#6B7280"))

        header_text = f"{config.APP_NAME} - 安全分析报告"
        canvas_obj.drawString(30, A4[1] - 30, header_text)

        logo_path = os.path.join(os.path.dirname(__file__), "logo.png")
        if os.path.exists(logo_path):
            try:
                canvas_obj.drawImage(logo_path, A4[0] - 80, A4[1] - 40, width=50, height=25,
                                     preserveAspectRatio=True)
            except Exception:
                pass

        page_num = canvas_obj.getPageNumber()
        footer_text = f"第 {page_num} 页"
        canvas_obj.drawCentredString(A4[0] / 2, 20, footer_text)
        canvas_obj.drawRightString(A4[0] - 30, 20, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        canvas_obj.restoreState()

    @with_read_session
    def generate_pdf_report(self, report_type: str, period_start: datetime, period_end: datetime,
                            stats_data: Optional[DailyStatsData] = None,
                            trend_data: Optional[TrendData] = None,
                            chart_paths: Optional[Dict[str, str]] = None,
                            compare_data: Optional[Dict[str, Any]] = None,
                            session: Optional[Session] = None) -> str:
        if session is None:
            raise ValueError("Session is required")

        report_type_enum = ReportTypeEnum(report_type.lower())
        report_date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_type_enum.value}_report_{report_date_str}.pdf"
        filepath = os.path.join(self.output_dir, filename)

        doc = SimpleDocTemplate(
            filepath,
            pagesize=A4,
            rightMargin=30,
            leftMargin=30,
            topMargin=50,
            bottomMargin=50
        )

        story = []

        story.extend(self._generate_cover_page(report_type_enum, period_start, period_end))
        story.append(PageBreak())
        story.extend(self._generate_toc())
        story.append(PageBreak())
        story.extend(self._generate_summary_section(stats_data, trend_data, compare_data))
        story.append(PageBreak())
        story.extend(self._generate_data_tables_section(stats_data))
        story.append(PageBreak())
        story.extend(self._generate_charts_section(chart_paths))
        story.append(PageBreak())
        story.extend(self._generate_analysis_section(stats_data, trend_data, compare_data))

        doc.build(story, onFirstPage=self._header_footer, onLaterPages=self._header_footer)

        log_with_context(logger, "info", f"PDF报表已生成: {filepath}",
                        report_type=report_type, filepath=filepath)

        return filepath

    def _generate_cover_page(self, report_type: ReportTypeEnum,
                              period_start: datetime, period_end: datetime) -> List[Any]:
        story = []

        type_names = {
            "daily": "日",
            "weekly": "周",
            "monthly": "月",
            "quarterly": "季度"
        }

        story.append(Spacer(1, 2 * inch))
        story.append(Paragraph(f"漏洞管理{type_names.get(report_type.value, '')}报告", self.title_style))
        story.append(Spacer(1, 0.5 * inch))

        period_text = f"统计周期: {period_start.strftime('%Y年%m月%d日')} - {period_end.strftime('%Y年%m月%d日')}"
        story.append(Paragraph(period_text, ParagraphStyle(
            "PeriodText", parent=self.body_style, fontSize=14, alignment=1, textColor=colors.HexColor("#6B7280")
        )))

        story.append(Spacer(1, 0.3 * inch))
        story.append(Paragraph(f"生成时间: {datetime.now().strftime('%Y年%m月%d日 %H:%M:%S')}",
                               ParagraphStyle("DateText", parent=self.body_style, fontSize=12, alignment=1)))

        story.append(Spacer(1, 2 * inch))
        story.append(Paragraph(config.APP_NAME, ParagraphStyle(
            "AppName", parent=self.body_style, fontSize=14, alignment=1,
            textColor=colors.HexColor("#1E40AF"), fontName="Helvetica-Bold"
        )))

        return story

    def _generate_toc(self) -> List[Any]:
        story = []
        story.append(Paragraph("目 录", self.title_style))
        story.append(Spacer(1, 0.3 * inch))

        toc_items = [
            ("一、报表概览", 3),
            ("二、数据统计表格", 4),
            ("三、可视化图表", 5),
            ("四、分析结论与建议", 6)
        ]

        for title, page in toc_items:
            story.append(Paragraph(f"{title}  {'·' * 50}  第 {page} 页", self.body_style))
            story.append(Spacer(1, 0.1 * inch))

        return story

    def _generate_summary_section(self, stats_data: Optional[DailyStatsData],
                                   trend_data: Optional[TrendData],
                                   compare_data: Optional[Dict[str, Any]]) -> List[Any]:
        story = []
        story.append(Paragraph("一、报表概览", self.heading2_style))

        if not stats_data:
            story.append(Paragraph("暂无统计数据", self.body_style))
            return story

        summary_data = [
            ["指标", "数值", "同比变化"],
            ["新增漏洞数", str(stats_data.new_vulns), self._format_change(compare_data, "new_vulns")],
            ["重复漏洞数", str(stats_data.repeated_vulns), self._format_change(compare_data, "repeated_vulns")],
            ["高危漏洞数", str(stats_data.high_risk_count), self._format_change(compare_data, "high_risk_count")],
            ["总工单数", str(stats_data.total_work_orders), self._format_change(compare_data, "total_work_orders")],
            ["修复率", f"{stats_data.fix_rate * 100:.1f}%", self._format_change(compare_data, "fix_rate", is_rate=True)],
            ["平均修复时长", f"{stats_data.avg_fix_duration_hours:.1f}小时", self._format_change(compare_data, "avg_fix_duration_hours", is_time=True)],
            ["超时率", f"{stats_data.timeout_rate * 100:.1f}%", self._format_change(compare_data, "timeout_rate", is_rate=True)],
            ["验证失败率", f"{stats_data.verify_fail_rate * 100:.1f}%", "-"],
            ["复盘任务数", str(stats_data.review_tasks_count), "-"]
        ]

        table = Table(summary_data, colWidths=[2 * inch, 1.5 * inch, 1.5 * inch])
        table.setStyle(self._get_table_style())
        story.append(table)
        story.append(Spacer(1, 0.3 * inch))

        if trend_data and trend_data.anomalies:
            story.append(Paragraph("⚠️ 异常检测", self.heading3_style))
            for period, value, atype in trend_data.anomalies:
                story.append(Paragraph(
                    f"• {period}: {trend_data.metric_name} 发生{atype}，当前值为 {value} {trend_data.metric_unit}",
                    self.highlight_style
                ))

        return story

    def _format_change(self, compare_data: Optional[Dict[str, Any]], key: str,
                       is_rate: bool = False, is_time: bool = False) -> str:
        if not compare_data or key not in compare_data:
            return "-"

        change = compare_data[key]
        if change is None:
            return "-"

        prefix = "+" if change > 0 else ""
        if is_rate:
            return f"{prefix}{change:.1f}%"
        else:
            return f"{prefix}{change:.1f}%"

    def _generate_data_tables_section(self, stats_data: Optional[DailyStatsData]) -> List[Any]:
        story = []
        story.append(Paragraph("二、数据统计表格", self.heading2_style))

        if not stats_data:
            story.append(Paragraph("暂无统计数据", self.body_style))
            return story

        story.append(Paragraph("2.1 漏洞严重等级分布", self.heading3_style))
        severity_data = [["严重等级", "数量", "占比"]]
        total = sum(stats_data.vulns_by_severity.values()) or 1
        for severity, count in stats_data.vulns_by_severity.items():
            severity_data.append([
                SEVERITY_NAMES.get(severity, severity),
                str(count),
                f"{count / total * 100:.1f}%"
            ])
        table = Table(severity_data, colWidths=[1.5 * inch, 1 * inch, 1 * inch])
        table.setStyle(self._get_table_style())
        story.append(table)
        story.append(Spacer(1, 0.2 * inch))

        story.append(Paragraph("2.2 资产类型分布", self.heading3_style))
        asset_data = [["资产类型", "数量", "占比"]]
        total = sum(stats_data.vulns_by_asset_type.values()) or 1
        for asset_type, count in sorted(stats_data.vulns_by_asset_type.items(), key=lambda x: -x[1]):
            asset_data.append([asset_type, str(count), f"{count / total * 100:.1f}%"])
        table = Table(asset_data, colWidths=[2 * inch, 1 * inch, 1 * inch])
        table.setStyle(self._get_table_style())
        story.append(table)
        story.append(Spacer(1, 0.2 * inch))

        story.append(Paragraph("2.3 部门分布", self.heading3_style))
        dept_data = [["部门", "漏洞数量", "占比"]]
        total = sum(stats_data.vulns_by_department.values()) or 1
        for dept, count in sorted(stats_data.vulns_by_department.items(), key=lambda x: -x[1]):
            dept_data.append([dept, str(count), f"{count / total * 100:.1f}%"])
        table = Table(dept_data, colWidths=[2 * inch, 1 * inch, 1 * inch])
        table.setStyle(self._get_table_style())
        story.append(table)
        story.append(Spacer(1, 0.2 * inch))

        story.append(Paragraph("2.4 工单状态分布", self.heading3_style))
        stage_data = [["状态", "数量", "占比"]]
        total = sum(stats_data.stage_distribution.values()) or 1
        for stage, count in stats_data.stage_distribution.items():
            stage_data.append([STAGE_NAMES.get(stage, stage), str(count), f"{count / total * 100:.1f}%"])
        table = Table(stage_data, colWidths=[1.5 * inch, 1 * inch, 1 * inch])
        table.setStyle(self._get_table_style())
        story.append(table)

        return story

    def _generate_charts_section(self, chart_paths: Optional[Dict[str, str]]) -> List[Any]:
        story = []
        story.append(Paragraph("三、可视化图表", self.heading2_style))

        if not chart_paths:
            story.append(Paragraph("暂无图表", self.body_style))
            return story

        chart_titles = {
            "severity_pie": "3.1 漏洞严重等级分布",
            "asset_type_pie": "3.2 资产类型分布",
            "daily_bar": "3.3 每日发现数量统计",
            "department_bar": "3.4 各部门漏洞数量对比",
            "stage_stacked_bar": "3.5 各阶段工单数量分布",
            "trend_line": "3.6 30天趋势分析",
            "fix_duration_trend": "3.7 各阶段平均修复时长",
            "heatmap": "3.8 漏洞类型×资产类型矩阵"
        }

        for key, path in chart_paths.items():
            if os.path.exists(path):
                story.append(Paragraph(chart_titles.get(key, key), self.heading3_style))
                try:
                    img = Image(path, width=6 * inch, height=4 * inch)
                    story.append(img)
                except Exception as e:
                    story.append(Paragraph(f"图片加载失败: {e}", self.body_style))
                story.append(Spacer(1, 0.2 * inch))

        return story

    def _generate_analysis_section(self, stats_data: Optional[DailyStatsData],
                                    trend_data: Optional[TrendData],
                                    compare_data: Optional[Dict[str, Any]]) -> List[Any]:
        story = []
        story.append(Paragraph("四、分析结论与建议", self.heading2_style))

        if not stats_data:
            story.append(Paragraph("暂无分析数据", self.body_style))
            return story

        conclusions = []

        if stats_data.high_risk_count > 0:
            conclusions.append((
                "安全风险",
                f"本期发现高危漏洞 {stats_data.high_risk_count} 个，建议优先处理，防止被利用。"
            ))

        if stats_data.timeout_rate > 0.1:
            conclusions.append((
                "效率问题",
                f"工单超时率为 {stats_data.timeout_rate * 100:.1f}%，超过10%警戒线，建议排查修复瓶颈。"
            ))

        if stats_data.verify_fail_rate > 0.15:
            conclusions.append((
                "质量问题",
                f"验证失败率为 {stats_data.verify_fail_rate * 100:.1f}%，建议加强修复质量管控。"
            ))

        if stats_data.fix_rate < 0.7:
            conclusions.append((
                "效率问题",
                f"修复率仅为 {stats_data.fix_rate * 100:.1f}%，建议提升修复效率。"
            ))

        if trend_data and trend_data.anomalies:
            for period, value, atype in trend_data.anomalies:
                if atype == "突增":
                    conclusions.append((
                        "异常告警",
                        f"{period} {trend_data.metric_name}{atype}至{value}{trend_data.metric_unit}，建议关注。"
                    ))

        if not conclusions:
            conclusions.append(("整体情况", "本期各项指标正常，继续保持良好的漏洞管理流程。"))

        for category, content in conclusions:
            story.append(Paragraph(f"• {category}: {content}", self.body_style))
            story.append(Spacer(1, 0.1 * inch))

        story.append(Spacer(1, 0.3 * inch))
        story.append(Paragraph("改进建议", self.heading3_style))

        suggestions = [
            "建议定期开展安全培训，提高开发人员安全意识",
            "建议引入自动化安全检测工具，实现左移安全",
            "建议优化漏洞修复流程，缩短修复周期",
            "建议建立漏洞管理KPI考核机制",
            "建议每季度进行一次全面的安全风险评估"
        ]

        suggestion_items = [ListItem(Paragraph(s, self.body_style)) for s in suggestions[:3]]
        story.append(ListFlowable(suggestion_items, bulletType="bullet"))

        return story

    def _get_table_style(self) -> TableStyle:
        return TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1E40AF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D1D5DB")),
            ("FONTSIZE", (0, 1), (-1, -1), 10),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#F9FAFB"), colors.white]),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("LEFTPADDING", (0, 0), (-1, -1), 8),
            ("RIGHTPADDING", (0, 0), (-1, -1), 8),
        ])


class ExcelReporter:
    def __init__(self, output_dir: Optional[str] = None):
        self.output_dir = output_dir or config.REPORT_OUTPUT_DIR
        self._ensure_output_dir()
        self._setup_styles()

    def _ensure_output_dir(self):
        os.makedirs(self.output_dir, exist_ok=True)

    def _setup_styles(self):
        self.header_font = Font(bold=True, size=11, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        self.center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        self.left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB")
        )
        self.high_risk_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        self.timeout_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
        self.stripe_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")

    @with_read_session
    def generate_excel_report(self, report_type: str, period_start: datetime, period_end: datetime,
                               stats_data: Optional[DailyStatsData] = None,
                               trend_data: Optional[TrendData] = None,
                               session: Optional[Session] = None) -> str:
        if session is None:
            raise ValueError("Session is required")

        report_type_enum = ReportTypeEnum(report_type.lower())
        report_date_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"{report_type_enum.value}_report_{report_date_str}.xlsx"
        filepath = os.path.join(self.output_dir, filename)

        wb = Workbook()

        self._create_vuln_detail_sheet(wb, period_start, period_end, session)
        self._create_summary_sheet(wb, stats_data, trend_data)

        if trend_data:
            self._create_trend_sheet(wb, trend_data)

        self._create_work_order_sheet(wb, period_start, period_end, session)

        wb.save(filepath)

        log_with_context(logger, "info", f"Excel报表已生成: {filepath}",
                        report_type=report_type, filepath=filepath)

        return filepath

    def _create_vuln_detail_sheet(self, wb: Workbook, period_start: datetime,
                                   period_end: datetime, session: Session):
        ws = wb.active
        ws.title = "漏洞明细"

        headers = ["ID", "漏洞标题", "CVE编号", "严重等级", "CVSS评分",
                   "资产名称", "资产IP", "资产类型", "所属部门", "发现时间",
                   "修复截止时间", "当前状态", "风险分数"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border

        vuln_instances = session.query(VulnerabilityInstance).join(Asset).join(Vulnerability).filter(
            and_(
                VulnerabilityInstance.discovery_time >= period_start,
                VulnerabilityInstance.discovery_time < period_end
            )
        ).all()

        for row, vi in enumerate(vuln_instances, 2):
            data = [
                vi.id,
                vi.vulnerability.title if vi.vulnerability else "",
                vi.vulnerability.cve_id if vi.vulnerability else "",
                SEVERITY_NAMES.get(vi.vulnerability.severity.value, vi.vulnerability.severity.value) if vi.vulnerability else "",
                float(vi.vulnerability.cvss_score) if vi.vulnerability and vi.vulnerability.cvss_score else 0,
                vi.asset.name if vi.asset else "",
                vi.asset.ip if vi.asset else "",
                vi.asset.type if vi.asset else "",
                vi.asset.department if vi.asset else "",
                vi.discovery_time.strftime("%Y-%m-%d %H:%M:%S"),
                vi.fix_deadline.strftime("%Y-%m-%d %H:%M:%S"),
                vi.fix_status.value,
                float(vi.risk_score)
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = self.center_align if col not in [2, 6] else self.left_align
                cell.border = self.border

                if row % 2 == 0:
                    cell.fill = self.stripe_fill

                severity = vi.vulnerability.severity.value if vi.vulnerability else ""
                if severity in ["critical", "high"] and col == 4:
                    cell.fill = self.high_risk_fill
                    cell.font = Font(bold=True, color="DC2626")

                if vi.fix_deadline < datetime.now() and vi.fix_status.value not in ["fixed", "verified"] and col == 11:
                    cell.fill = self.timeout_fill

        self._auto_adjust_column_width(ws)

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(vuln_instances) + 1}"
        ws.freeze_panes = "A2"

    def _create_summary_sheet(self, wb: Workbook, stats_data: Optional[DailyStatsData],
                               trend_data: Optional[TrendData]):
        ws = wb.create_sheet("统计汇总")

        if not stats_data:
            ws["A1"] = "暂无统计数据"
            return

        row = 1

        ws.cell(row=row, column=1, value="一、核心指标")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1E40AF")
        ws.merge_cells(f"A{row}:D{row}")
        row += 2

        summary_headers = ["指标", "数值", "计算公式", "备注"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1

        metrics = [
            ("新增漏洞数", stats_data.new_vulns, f"=COUNTIF(漏洞明细!A:A,\">0\")", "统计周期内新发现"),
            ("重复漏洞数", stats_data.repeated_vulns, "=B2-B3", "历史曾发现过"),
            ("高危漏洞数", stats_data.high_risk_count, '=COUNTIF(漏洞明细!D:D,"严重")+COUNTIF(漏洞明细!D:D,"高危")', "严重+高危"),
            ("总工单数", stats_data.total_work_orders, "=COUNTA(工单明细!A:A)-1", "周期内创建"),
            ("已关闭工单数", stats_data.closed_count, '=COUNTIF(工单明细!D:D,"已关闭")', "状态为已关闭"),
            ("修复率", f"{stats_data.fix_rate * 100:.1f}%", "=B6/B5", "已关闭/总"),
            ("平均修复时长", f"{stats_data.avg_fix_duration_hours:.1f}小时", "", "创建到关闭"),
            ("超时工单数", stats_data.timeout_count, "", "超过截止时间"),
            ("超时率", f"{stats_data.timeout_rate * 100:.1f}%", "=B9/B5", "超时/总工单"),
            ("验证失败率", f"{stats_data.verify_fail_rate * 100:.1f}%", "", "验证不通过比例")
        ]

        for metric, value, formula, note in metrics:
            ws.cell(row=row, column=1, value=metric).border = self.border
            cell = ws.cell(row=row, column=2, value=value)
            cell.border = self.border
            ws.cell(row=row, column=3, value=formula).border = self.border
            ws.cell(row=row, column=4, value=note).border = self.border
            for col in range(1, 5):
                ws.cell(row=row, column=col).alignment = self.center_align
                if row % 2 == 0:
                    ws.cell(row=row, column=col).fill = self.stripe_fill
            row += 1

        row += 2
        ws.cell(row=row, column=1, value="二、漏洞严重等级分布")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1E40AF")
        ws.merge_cells(f"A{row}:D{row}")
        row += 2

        severity_headers = ["严重等级", "数量", "占比"]
        for col, header in enumerate(severity_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1

        total = sum(stats_data.vulns_by_severity.values()) or 1
        for severity, count in stats_data.vulns_by_severity.items():
            ws.cell(row=row, column=1, value=SEVERITY_NAMES.get(severity, severity)).border = self.border
            ws.cell(row=row, column=2, value=count).border = self.border
            pct_cell = ws.cell(row=row, column=3, value=f"{count/total*100:.1f}%")
            pct_cell.border = self.border
            for col in range(1, 4):
                ws.cell(row=row, column=col).alignment = self.center_align
                if row % 2 == 0:
                    ws.cell(row=row, column=col).fill = self.stripe_fill
            if severity in ["critical", "high"]:
                ws.cell(row=row, column=1).fill = self.high_risk_fill
            row += 1

        chart_start_row = row
        pie = PieChart()
        labels = Reference(ws, min_col=1, min_row=chart_start_row - len(stats_data.vulns_by_severity), max_row=chart_start_row - 1)
        data = Reference(ws, min_col=2, min_row=chart_start_row - len(stats_data.vulns_by_severity) - 1, max_row=chart_start_row - 1)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "漏洞严重等级分布"
        ws.add_chart(pie, f"F{chart_start_row - 3}")

        row += 5
        ws.cell(row=row, column=1, value="三、部门分布统计")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1E40AF")
        ws.merge_cells(f"A{row}:D{row}")
        row += 2

        dept_headers = ["部门", "漏洞数量", "占比"]
        for col, header in enumerate(dept_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1

        dept_total = sum(stats_data.vulns_by_department.values()) or 1
        for dept, count in sorted(stats_data.vulns_by_department.items(), key=lambda x: -x[1]):
            ws.cell(row=row, column=1, value=dept).border = self.border
            ws.cell(row=row, column=2, value=count).border = self.border
            ws.cell(row=row, column=3, value=f"{count/dept_total*100:.1f}%").border = self.border
            for col in range(1, 4):
                ws.cell(row=row, column=col).alignment = self.center_align
                if row % 2 == 0:
                    ws.cell(row=row, column=col).fill = self.stripe_fill
            row += 1

        self._auto_adjust_column_width(ws)

    def _create_trend_sheet(self, wb: Workbook, trend_data: TrendData):
        ws = wb.create_sheet("趋势数据")

        row = 1
        ws.cell(row=row, column=1, value=f"{trend_data.metric_name} 趋势数据")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14, color="1E40AF")
        ws.merge_cells(f"A{row}:E{row}")
        row += 2

        ws.cell(row=row, column=1, value=f"统计周期: {trend_data.period}")
        row += 1
        ws.cell(row=row, column=1, value=f"环比变化: {'+' if trend_data.mom_change > 0 else ''}{trend_data.mom_change:.1f}%")
        ws.cell(row=row, column=1).font = Font(color="DC2626" if trend_data.mom_change > 0 else "16A34A", bold=True)
        row += 2

        headers = ["周期", "本期值", "上期值", "环比变化", "预测值"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border
        row += 1

        max_len = max(len(trend_data.periods), len(trend_data.forecast))
        for i in range(max_len):
            period = trend_data.periods[i] if i < len(trend_data.periods) else f"预测{i - len(trend_data.periods) + 1}"
            curr_val = trend_data.values[i] if i < len(trend_data.values) else ""
            last_val = trend_data.values_last_period[i] if i < len(trend_data.values_last_period) else ""

            if last_val and curr_val and last_val != 0:
                change = ((curr_val - last_val) / last_val) * 100
                change_str = f"{'+' if change > 0 else ''}{change:.1f}%"
            else:
                change_str = ""

            forecast_val = trend_data.forecast[i - len(trend_data.values)] if i >= len(trend_data.values) else ""

            ws.cell(row=row, column=1, value=period).border = self.border
            ws.cell(row=row, column=2, value=curr_val).border = self.border
            ws.cell(row=row, column=3, value=last_val).border = self.border
            change_cell = ws.cell(row=row, column=4, value=change_str)
            change_cell.border = self.border
            if "预警" not in period and change_str:
                change_cell.font = Font(color="DC2626" if change > 0 else "16A34A")
            ws.cell(row=row, column=5, value=forecast_val).border = self.border

            for col in range(1, 6):
                ws.cell(row=row, column=col).alignment = self.center_align
                if row % 2 == 0:
                    ws.cell(row=row, column=col).fill = self.stripe_fill
            row += 1

        if trend_data.anomalies:
            row += 1
            ws.cell(row=row, column=1, value="⚠️ 异常检测")
            ws.cell(row=row, column=1).font = Font(bold=True, color="DC2626")
            row += 1
            for period, value, atype in trend_data.anomalies:
                ws.cell(row=row, column=1, value=f"{period}: {trend_data.metric_name}{atype}至{value}{trend_data.metric_unit}")
                ws.cell(row=row, column=1).font = Font(color="DC2626")
                row += 1

        chart = LineChart()
        chart.title = f"{trend_data.metric_name} 趋势图"
        chart.style = 13
        chart.y_axis.title = f"{trend_data.metric_name} ({trend_data.metric_unit})"
        chart.x_axis.title = "周期"

        data = Reference(ws, min_col=2, min_row=row - max_len - 1, max_col=3, max_row=row - 1)
        cats = Reference(ws, min_col=1, min_row=row - max_len, max_row=row - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws.add_chart(chart, f"G{row - max_len - 2}")

        self._auto_adjust_column_width(ws)

    def _create_work_order_sheet(self, wb: Workbook, period_start: datetime,
                                  period_end: datetime, session: Session):
        ws = wb.create_sheet("工单明细")

        headers = ["工单号", "漏洞ID", "指派人", "当前状态", "创建时间",
                   "开始修复时间", "修复完成时间", "验证时间", "关闭时间",
                   "截止时间", "优先级", "是否超时", "升级等级"]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.center_align
            cell.border = self.border

        work_orders = session.query(WorkOrder).filter(
            and_(
                WorkOrder.created_at >= period_start,
                WorkOrder.created_at < period_end
            )
        ).all()

        for row, wo in enumerate(work_orders, 2):
            is_timeout = "是" if (wo.deadline < datetime.now() and wo.status != WorkOrderStatusEnum.CLOSED) else "否"
            data = [
                wo.id,
                wo.vuln_instance_id,
                wo.assignee,
                STAGE_NAMES.get(wo.status.value, wo.status.value),
                wo.created_at.strftime("%Y-%m-%d %H:%M:%S") if wo.created_at else "",
                wo.started_at.strftime("%Y-%m-%d %H:%M:%S") if wo.started_at else "",
                wo.fixed_at.strftime("%Y-%m-%d %H:%M:%S") if wo.fixed_at else "",
                wo.verified_at.strftime("%Y-%m-%d %H:%M:%S") if wo.verified_at else "",
                wo.closed_at.strftime("%Y-%m-%d %H:%M:%S") if wo.closed_at else "",
                wo.deadline.strftime("%Y-%m-%d %H:%M:%S") if wo.deadline else "",
                wo.priority if wo.priority else "",
                is_timeout,
                wo.escalation_level
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.alignment = self.center_align
                cell.border = self.border

                if row % 2 == 0:
                    cell.fill = self.stripe_fill

                if col == 4 and wo.status.value in ["pending", "fixing"]:
                    cell.font = Font(color="DC2626", bold=True)

                if col == 12 and is_timeout == "是":
                    cell.fill = self.timeout_fill
                    cell.font = Font(color="DC2626", bold=True)

                if col == 13 and wo.escalation_level >= 2:
                    cell.fill = self.high_risk_fill
                    cell.font = Font(color="DC2626", bold=True)

        self._auto_adjust_column_width(ws)

        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(work_orders) + 1}"
        ws.freeze_panes = "A2"

        red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        yellow_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        ws.conditional_formatting.add(
            f"L2:L{len(work_orders) + 1}",
            CellIsRule(operator="equal", formula=['"是"'], fill=yellow_fill, font=Font(color="DC2626", bold=True))
        )

        ws.conditional_formatting.add(
            f"M2:M{len(work_orders) + 1}",
            CellIsRule(operator="greaterThanOrEqual", formula=["2"], fill=red_fill, font=Font(color="DC2626", bold=True))
        )

    def _auto_adjust_column_width(self, ws):
        for column_cells in ws.columns:
            length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            adjusted_width = min(max(length + 2, 10), 50)
            ws.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width


class ReportManager:
    def __init__(self):
        self.stats_engine = StatsEngine()
        self.trend_analyzer = TrendAnalyzer()
        self.chart_generator = ChartGenerator()
        self.pdf_reporter = PdfReporter()
        self.excel_reporter = ExcelReporter()
        self._ensure_directories()

    def _ensure_directories(self):
        os.makedirs(config.REPORT_OUTPUT_DIR, exist_ok=True)
        os.makedirs(os.path.join(config.REPORT_OUTPUT_DIR, "daily"), exist_ok=True)
        os.makedirs(os.path.join(config.REPORT_OUTPUT_DIR, "weekly"), exist_ok=True)
        os.makedirs(os.path.join(config.REPORT_OUTPUT_DIR, "monthly"), exist_ok=True)
        os.makedirs(os.path.join(config.REPORT_OUTPUT_DIR, "charts"), exist_ok=True)

    def _get_report_dir(self, report_type: str) -> str:
        base_dir = config.REPORT_OUTPUT_DIR
        type_dirs = {
            "daily": "daily",
            "weekly": "weekly",
            "monthly": "monthly",
            "quarterly": "quarterly"
        }
        return os.path.join(base_dir, type_dirs.get(report_type.lower(), ""))

    def _get_compare_data(self, current_stats: DailyStatsData,
                           prev_period_start: datetime,
                           prev_period_end: datetime) -> Dict[str, float]:
        compare_data = {}
        try:
            prev_stats = self.stats_engine.calculate_custom_stats(
                prev_period_start, prev_period_end
            )

            compare_data["new_vulns"] = self._calc_change(
                current_stats.new_vulns, prev_stats.get("total_vulns", 0)
            )
            compare_data["total_work_orders"] = self._calc_change(
                current_stats.total_work_orders, prev_stats.get("total_work_orders", 0)
            )
            compare_data["fix_rate"] = self._calc_change(
                current_stats.fix_rate, prev_stats.get("fix_rate", 0)
            )
        except Exception as e:
            log_with_context(logger, "warning", f"计算对比数据失败: {e}")

        return compare_data

    def _calc_change(self, current: float, previous: float) -> float:
        if previous == 0:
            return 100.0 if current > 0 else 0.0
        return ((current - previous) / previous) * 100

    @with_session
    def run_daily_report(self, report_date: Optional[datetime] = None,
                          auto_send: bool = True,
                          session: Optional[Session] = None) -> Dict[str, Any]:
        if session is None:
            raise ValueError("Session is required")

        report_date = report_date or datetime.now() - timedelta(days=1)
        report_date = datetime(report_date.year, report_date.month, report_date.day)
        period_end = report_date + timedelta(days=1)

        log_with_context(logger, "info", f"开始生成日报: {report_date.strftime('%Y-%m-%d')}")

        try:
            output_dir = os.path.join(self._get_report_dir("daily"),
                                      report_date.strftime("%Y%m%d"))
            os.makedirs(output_dir, exist_ok=True)

            self.chart_generator.output_dir = output_dir
            self.pdf_reporter.output_dir = output_dir
            self.excel_reporter.output_dir = output_dir

            stats_data = self.stats_engine.calculate_daily_stats(report_date)

            trend_start = report_date - timedelta(days=30)
            trend_data = self.trend_analyzer.generate_trend_data(
                trend_start, period_end, granularity="day", metric="new_vulns"
            )

            chart_paths = self.chart_generator.generate_charts(stats_data, trend_data)

            prev_start = report_date - timedelta(days=1)
            prev_end = report_date
            compare_data = self._get_compare_data(stats_data, prev_start, prev_end)

            pdf_path = self.pdf_reporter.generate_pdf_report(
                "daily", report_date, period_end,
                stats_data=stats_data, trend_data=trend_data,
                chart_paths=chart_paths, compare_data=compare_data
            )

            excel_path = self.excel_reporter.generate_excel_report(
                "daily", report_date, period_end,
                stats_data=stats_data, trend_data=trend_data
            )

            report_record = Report(
                type=ReportTypeEnum.DAILY,
                period_start=report_date,
                period_end=period_end,
                generated_by="system",
                file_path_pdf=pdf_path,
                file_path_excel=excel_path,
                summary_stats=stats_data.to_dict()
            )
            session.add(report_record)
            session.flush()

            result = {
                "success": True,
                "report_id": report_record.id,
                "report_type": "daily",
                "period": report_date.strftime("%Y-%m-%d"),
                "pdf_path": pdf_path,
                "excel_path": excel_path,
                "stats": stats_data.to_dict(),
                "chart_paths": chart_paths
            }

            if auto_send:
                recipients = config.notification.__dict__.get("MANAGEMENT_EMAILS", [])
                if recipients:
                    send_result = self.send_report(result, recipients)
                    result["email_sent"] = send_result["success"]

            log_with_context(logger, "info", f"日报生成完成: {report_date.strftime('%Y-%m-%d')}",
                            report_id=report_record.id)

            return result

        except Exception as e:
            log_with_context(logger, "error", f"日报生成失败: {e}",
                            report_date=report_date.strftime("%Y-%m-%d"))
            raise

    @with_session
    def run_weekly_report(self, week_end: Optional[datetime] = None,
                           auto_send: bool = True,
                           session: Optional[Session] = None) -> Dict[str, Any]:
        if session is None:
            raise ValueError("Session is required")

        week_end = week_end or datetime.now()
        week_end = datetime(week_end.year, week_end.month, week_end.day)
        week_start = week_end - timedelta(days=6)

        log_with_context(logger, "info", f"开始生成周报: {week_start.strftime('%Y-%m-%d')} ~ {week_end.strftime('%Y-%m-%d')}")

        try:
            output_dir = os.path.join(
                self._get_report_dir("weekly"),
                f"{week_start.strftime('%Y%m%d')}_{week_end.strftime('%Y%m%d')}"
            )
            os.makedirs(output_dir, exist_ok=True)

            self.chart_generator.output_dir = output_dir
            self.pdf_reporter.output_dir = output_dir
            self.excel_reporter.output_dir = output_dir

            daily_stats = []
            for i in range(7):
                day = week_start + timedelta(days=i)
                daily_stats.append(self.stats_engine.calculate_daily_stats(day))

            stats_data = self._aggregate_daily_stats(daily_stats)

            trend_start = week_start - timedelta(weeks=4)
            trend_data = self.trend_analyzer.generate_trend_data(
                trend_start, week_end + timedelta(days=1),
                granularity="week", metric="new_vulns"
            )

            chart_paths = self.chart_generator.generate_charts(stats_data, trend_data)

            prev_week_start = week_start - timedelta(weeks=1)
            prev_week_end = week_start - timedelta(days=1)
            compare_data = self._get_compare_data(stats_data, prev_week_start, prev_week_end)

            pdf_path = self.pdf_reporter.generate_pdf_report(
                "weekly", week_start, week_end,
                stats_data=stats_data, trend_data=trend_data,
                chart_paths=chart_paths, compare_data=compare_data
            )

            excel_path = self.excel_reporter.generate_excel_report(
                "weekly", week_start, week_end,
                stats_data=stats_data, trend_data=trend_data
            )

            report_record = Report(
                type=ReportTypeEnum.WEEKLY,
                period_start=week_start,
                period_end=week_end,
                generated_by="system",
                file_path_pdf=pdf_path,
                file_path_excel=excel_path,
                summary_stats=stats_data.to_dict()
            )
            session.add(report_record)
            session.flush()

            result = {
                "success": True,
                "report_id": report_record.id,
                "report_type": "weekly",
                "period": f"{week_start.strftime('%Y-%m-%d')} ~ {week_end.strftime('%Y-%m-%d')}",
                "pdf_path": pdf_path,
                "excel_path": excel_path,
                "stats": stats_data.to_dict(),
                "chart_paths": chart_paths,
                "weekly_comparison": compare_data
            }

            if auto_send:
                recipients = config.notification.__dict__.get("MANAGEMENT_EMAILS", [])
                if recipients:
                    send_result = self.send_report(result, recipients)
                    result["email_sent"] = send_result["success"]

            log_with_context(logger, "info", f"周报生成完成", report_id=report_record.id)

            return result

        except Exception as e:
            log_with_context(logger, "error", f"周报生成失败: {e}")
            raise

    def _aggregate_daily_stats(self, daily_stats: List[DailyStatsData]) -> DailyStatsData:
        if not daily_stats:
            return DailyStatsData(date="aggregate")

        first = daily_stats[0]
        aggregated = DailyStatsData(
            date=f"{first.date} ~ {daily_stats[-1].date}"
        )

        for stats in daily_stats:
            aggregated.total_vulns += stats.total_vulns
            aggregated.new_vulns += stats.new_vulns
            aggregated.repeated_vulns += stats.repeated_vulns
            aggregated.high_risk_count += stats.high_risk_count
            aggregated.total_work_orders += stats.total_work_orders
            aggregated.pending_count += stats.pending_count
            aggregated.fixing_count += stats.fixing_count
            aggregated.fixed_count += stats.fixed_count
            aggregated.verifying_count += stats.verifying_count
            aggregated.closed_count += stats.closed_count
            aggregated.timeout_count += stats.timeout_count
            aggregated.verify_fail_count += stats.verify_fail_count
            aggregated.review_tasks_count += stats.review_tasks_count

            for key, value in stats.vulns_by_severity.items():
                aggregated.vulns_by_severity[key] = aggregated.vulns_by_severity.get(key, 0) + value

            for key, value in stats.vulns_by_asset_type.items():
                aggregated.vulns_by_asset_type[key] = aggregated.vulns_by_asset_type.get(key, 0) + value

            for key, value in stats.vulns_by_department.items():
                aggregated.vulns_by_department[key] = aggregated.vulns_by_department.get(key, 0) + value

            for key, value in stats.stage_distribution.items():
                aggregated.stage_distribution[key] = aggregated.stage_distribution.get(key, 0) + value

        if aggregated.closed_count > 0:
            total_duration = sum(s.avg_fix_duration_hours * s.closed_count for s in daily_stats if s.closed_count > 0)
            aggregated.avg_fix_duration_hours = total_duration / aggregated.closed_count

        if aggregated.total_work_orders > 0:
            aggregated.fix_rate = aggregated.closed_count / aggregated.total_work_orders
            aggregated.timeout_rate = aggregated.timeout_count / aggregated.total_work_orders

        total_verify = sum(s.verify_fail_count + (s.total_work_orders - s.verify_fail_count) for s in daily_stats)
        if total_verify > 0:
            aggregated.verify_fail_rate = aggregated.verify_fail_count / total_verify

        for stage in ["pending", "fixing", "fixed", "verifying", "closed"]:
            durations = [s.avg_stage_durations.get(stage, 0) for s in daily_stats if s.avg_stage_durations.get(stage, 0) > 0]
            if durations:
                aggregated.avg_stage_durations[stage] = sum(durations) / len(durations)

        return aggregated

    def send_report(self, report: Dict[str, Any], recipients: List[str]) -> Dict[str, Any]:
        if not recipients:
            return {"success": False, "error": "未指定收件人"}

        try:
            smtp_config = config.notification
            if not smtp_config.SMTP_HOST or not smtp_config.SMTP_USERNAME:
                log_with_context(logger, "warning", "邮件配置不完整，跳过发送")
                return {"success": False, "error": "邮件配置不完整"}

            msg = MIMEMultipart()
            msg["From"] = smtp_config.SMTP_FROM
            msg["To"] = ", ".join(recipients)
            msg["Subject"] = f"【{config.APP_NAME}】{report.get('period', '')} {report.get('report_type', '')}报告"

            body = self._generate_email_body(report)
            msg.attach(MIMEText(body, "html", "utf-8"))

            for path_key in ["pdf_path", "excel_path"]:
                file_path = report.get(path_key)
                if file_path and os.path.exists(file_path):
                    self._attach_file(msg, file_path)

            with smtplib.SMTP(smtp_config.SMTP_HOST, smtp_config.SMTP_PORT) as server:
                if smtp_config.SMTP_USE_TLS:
                    server.starttls()
                server.login(smtp_config.SMTP_USERNAME, smtp_config.SMTP_PASSWORD)
                server.send_message(msg)

            log_with_context(logger, "info", f"报表邮件已发送给 {len(recipients)} 位收件人",
                            recipients=recipients)

            return {"success": True, "recipients": recipients}

        except Exception as e:
            log_with_context(logger, "error", f"报表邮件发送失败: {e}")
            return {"success": False, "error": str(e)}

    def _generate_email_body(self, report: Dict[str, Any]) -> str:
        stats = report.get("stats", {})
        period = report.get("period", "")
        report_type = report.get("report_type", "")

        type_names = {"daily": "日", "weekly": "周", "monthly": "月", "quarterly": "季度"}
        type_name = type_names.get(report_type, "")

        body = f"""
        <html>
        <head>
            <meta charset="utf-8">
            <style>
                body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                .header {{ background: #1E40AF; color: white; padding: 20px; text-align: center; }}
                .content {{ padding: 20px; }}
                .stats-table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
                .stats-table th, .stats-table td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
                .stats-table th {{ background: #f5f5f5; font-weight: bold; }}
                .highlight {{ color: #DC2626; font-weight: bold; }}
                .success {{ color: #16A34A; }}
                .warning {{ color: #CA8A04; }}
                .footer {{ margin-top: 30px; padding: 20px; border-top: 1px solid #eee; color: #666; font-size: 12px; }}
            </style>
        </head>
        <body>
            <div class="header">
                <h2>{config.APP_NAME} - 漏洞管理{type_name}报告</h2>
                <p>统计周期: {period}</p>
            </div>
            <div class="content">
                <h3>一、核心指标概览</h3>
                <table class="stats-table">
                    <tr><th>指标</th><th>数值</th></tr>
                    <tr><td>新增漏洞数</td><td class="{'highlight' if stats.get('new_vulns', 0) > 0 else ''}">{stats.get('new_vulns', 0)}</td></tr>
                    <tr><td>高危漏洞数</td><td class="highlight">{stats.get('high_risk_count', 0)}</td></tr>
                    <tr><td>总工单数</td><td>{stats.get('total_work_orders', 0)}</td></tr>
                    <tr><td>修复率</td><td class="{'success' if stats.get('fix_rate', 0) >= 0.8 else 'warning'}">{stats.get('fix_rate', 0) * 100:.1f}%</td></tr>
                    <tr><td>平均修复时长</td><td>{stats.get('avg_fix_duration_hours', 0):.1f}小时</td></tr>
                    <tr><td>超时率</td><td class="{'warning' if stats.get('timeout_rate', 0) > 0.1 else ''}">{stats.get('timeout_rate', 0) * 100:.1f}%</td></tr>
                </table>
        """

        if report.get("weekly_comparison"):
            body += """
                <h3>二、环比变化</h3>
                <table class="stats-table">
                    <tr><th>指标</th><th>环比变化</th></tr>
            """
            for key, value in report["weekly_comparison"].items():
                key_names = {"new_vulns": "新增漏洞", "total_work_orders": "总工单", "fix_rate": "修复率"}
                color = "highlight" if value > 0 and key != "fix_rate" else ("success" if key == "fix_rate" and value > 0 else "")
                body += f'<tr><td>{key_names.get(key, key)}</td><td class="{color}">{"+" if value > 0 else ""}{value:.1f}%</td></tr>'
            body += "</table>"

        anomalies = report.get("stats", {}).get("anomalies", [])
        if anomalies:
            body += """
                <h3>⚠️ 异常提醒</h3>
                <ul>
            """
            for anomaly in anomalies:
                body += f'<li class="highlight">{anomaly}</li>'
            body += "</ul>"

        body += f"""
                <p>详细数据请查看附件中的 PDF 和 Excel 报表。</p>
            </div>
            <div class="footer">
                <p>此邮件由 {config.APP_NAME} 自动发送，请勿直接回复。</p>
                <p>生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
        </body>
        </html>
        """
        return body

    def _attach_file(self, msg: MIMEMultipart, file_path: str):
        with open(file_path, "rb") as f:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        filename = os.path.basename(file_path)
        part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
        msg.attach(part)

    @with_read_session
    def query_reports(self, report_type: Optional[str] = None,
                       start_date: Optional[datetime] = None,
                       end_date: Optional[datetime] = None,
                       session: Optional[Session] = None) -> List[Dict[str, Any]]:
        if session is None:
            raise ValueError("Session is required")

        query = session.query(Report)

        if report_type:
            query = query.filter(Report.type == ReportTypeEnum(report_type.lower()))

        if start_date:
            query = query.filter(Report.period_start >= start_date)

        if end_date:
            query = query.filter(Report.period_end <= end_date)

        reports = query.order_by(Report.created_at.desc()).all()

        return [
            {
                "id": r.id,
                "type": r.type.value,
                "period_start": r.period_start.strftime("%Y-%m-%d"),
                "period_end": r.period_end.strftime("%Y-%m-%d"),
                "generated_at": r.generated_at.strftime("%Y-%m-%d %H:%M:%S"),
                "generated_by": r.generated_by,
                "pdf_path": r.file_path_pdf,
                "excel_path": r.file_path_excel,
                "summary_stats": r.summary_stats
            }
            for r in reports
        ]

    def generate_custom_report(self, report_type: str,
                                period_start: datetime,
                                period_end: datetime,
                                output_dir: Optional[str] = None) -> Dict[str, Any]:
        log_with_context(logger, "info", f"开始生成自定义报表: {period_start} ~ {period_end}")

        if output_dir:
            self.chart_generator.output_dir = output_dir
            self.pdf_reporter.output_dir = output_dir
            self.excel_reporter.output_dir = output_dir

        daily_stats = []
        current = period_start
        while current < period_end:
            daily_stats.append(self.stats_engine.calculate_daily_stats(current))
            current += timedelta(days=1)

        stats_data = self._aggregate_daily_stats(daily_stats)

        trend_data = self.trend_analyzer.generate_trend_data(
            period_start - timedelta(days=30), period_end,
            granularity="day", metric="new_vulns"
        )

        chart_paths = self.chart_generator.generate_charts(stats_data, trend_data)

        prev_start = period_start - (period_end - period_start)
        prev_end = period_start
        compare_data = self._get_compare_data(stats_data, prev_start, prev_end)

        pdf_path = self.pdf_reporter.generate_pdf_report(
            report_type, period_start, period_end,
            stats_data=stats_data, trend_data=trend_data,
            chart_paths=chart_paths, compare_data=compare_data
        )

        excel_path = self.excel_reporter.generate_excel_report(
            report_type, period_start, period_end,
            stats_data=stats_data, trend_data=trend_data
        )

        result = {
            "success": True,
            "report_type": report_type,
            "period": f"{period_start.strftime('%Y-%m-%d')} ~ {period_end.strftime('%Y-%m-%d')}",
            "pdf_path": pdf_path,
            "excel_path": excel_path,
            "stats": stats_data.to_dict(),
            "chart_paths": chart_paths
        }

        log_with_context(logger, "info", "自定义报表生成完成")

        return result