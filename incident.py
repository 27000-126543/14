import os
import io
import csv
import json
import zipfile
import enum
import asyncio
from datetime import datetime, timedelta, timezone
from dataclasses import dataclass, field, asdict
from typing import List, Optional, Dict, Any, Tuple, Union
from decimal import Decimal

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import and_, or_, func, desc, asc, text
from sqlalchemy.orm import Session

from config import config
from database import db_manager, with_session, with_read_session
from logger import logger, log_audit, with_log_context
from models import (
    Base,
    SeverityEnum,
    Asset,
    Vulnerability,
    VulnerabilityInstance,
    WorkOrder,
    WorkOrderStatusEnum,
    Notification,
    AuditLog,
    EscalationRecord,
    VerificationRecord,
    ReviewTask,
    ResponsePlan,
    Incident,
    IncidentTimeline,
)


class IncidentType(str, enum.Enum):
    DATA_BREACH = "data_breach"
    INTRUSION = "intrusion"
    RANSOMWARE = "ransomware"
    DDOS = "ddos"
    COMPLIANCE_VIOLATION = "compliance_violation"
    OTHER = "other"


class IncidentStatus(str, enum.Enum):
    OPEN = "open"
    INVESTIGATING = "investigating"
    MITIGATED = "mitigated"
    RESOLVED = "resolved"
    CLOSED = "closed"


class TimelineEventType(str, enum.Enum):
    DETECTED = "detected"
    NOTIFIED = "notified"
    CONTAINED = "contained"
    ERADICATED = "eradicated"
    RECOVERED = "recovered"
    COMMUNICATED = "communicated"
    OTHER = "other"


class ExportFormat(str, enum.Enum):
    CSV = "csv"
    EXCEL = "excel"
    JSON = "json"


class QueryType(str, enum.Enum):
    VULNERABILITIES = "vulnerabilities"
    WORK_ORDERS = "work_orders"
    INCIDENTS = "incidents"
    LIFECYCLE = "lifecycle"


class ExportStatus(str, enum.Enum):
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"


@dataclass
class QueryFilter:
    vuln_id: Optional[int] = None
    cve_id: Optional[str] = None
    asset_name: Optional[str] = None
    asset_ip: Optional[str] = None
    department: Optional[str] = None
    importance: Optional[int] = None
    discovery_time_start: Optional[datetime] = None
    discovery_time_end: Optional[datetime] = None
    fix_time_start: Optional[datetime] = None
    fix_time_end: Optional[datetime] = None
    close_time_start: Optional[datetime] = None
    close_time_end: Optional[datetime] = None
    work_order_status: Optional[WorkOrderStatusEnum] = None
    risk_level: Optional[SeverityEnum] = None
    severity: Optional[SeverityEnum] = None
    operator: Optional[str] = None
    assignee: Optional[str] = None
    incident_type: Optional[IncidentType] = None
    incident_status: Optional[IncidentStatus] = None
    custom_filters: Optional[Dict[str, Any]] = None


@dataclass
class TimelineEvent:
    event_type: TimelineEventType
    description: str
    operator: str
    timestamp: datetime


@dataclass
class IncidentReport:
    incident_id: int
    title: str
    incident_type: IncidentType
    severity: SeverityEnum
    status: IncidentStatus
    timeline: List[TimelineEvent]
    affected_assets: List[Dict[str, Any]]
    affected_users_count: int
    related_vulnerabilities: List[Dict[str, Any]]
    response_measures: List[str]
    root_cause: str
    improvement_suggestions: List[str]
    mtta: Optional[float] = None
    mttr: Optional[float] = None
    business_impact: str = ""
    generated_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))

    def to_dict(self) -> Dict[str, Any]:
        return {
            "incident_id": self.incident_id,
            "title": self.title,
            "incident_type": self.incident_type.value,
            "severity": self.severity.value,
            "status": self.status.value,
            "timeline": [
                {
                    "event_type": e.event_type.value,
                    "description": e.description,
                    "operator": e.operator,
                    "timestamp": e.timestamp.isoformat()
                }
                for e in self.timeline
            ],
            "affected_assets": self.affected_assets,
            "affected_users_count": self.affected_users_count,
            "related_vulnerabilities": self.related_vulnerabilities,
            "response_measures": self.response_measures,
            "root_cause": self.root_cause,
            "improvement_suggestions": self.improvement_suggestions,
            "mtta": self.mtta,
            "mttr": self.mttr,
            "business_impact": self.business_impact,
            "generated_at": self.generated_at.isoformat()
        }


@dataclass
class ExportTask:
    task_id: str
    query_type: QueryType
    export_format: ExportFormat
    filters: Dict[str, Any]
    fields: List[str]
    operator: str
    status: ExportStatus = ExportStatus.PENDING
    file_path: Optional[str] = None
    total_records: int = 0
    exported_records: int = 0
    error_message: Optional[str] = None
    created_at: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    completed_at: Optional[datetime] = None

    def to_dict(self) -> Dict[str, Any]:
        return {
            "task_id": self.task_id,
            "query_type": self.query_type.value,
            "export_format": self.export_format.value,
            "filters": self.filters,
            "fields": self.fields,
            "operator": self.operator,
            "status": self.status.value,
            "file_path": self.file_path,
            "total_records": self.total_records,
            "exported_records": self.exported_records,
            "error_message": self.error_message,
            "created_at": self.created_at.isoformat(),
            "completed_at": self.completed_at.isoformat() if self.completed_at else None
        }


class TimelineManager:
    def __init__(self):
        self.logger = logger

    @staticmethod
    def _map_event_type(event_type: TimelineEventType) -> str:
        mapping = {
            TimelineEventType.DETECTED: "detected",
            TimelineEventType.NOTIFIED: "comment",
            TimelineEventType.CONTAINED: "mitigated",
            TimelineEventType.ERADICATED: "mitigated",
            TimelineEventType.RECOVERED: "resolved",
            TimelineEventType.COMMUNICATED: "comment",
            TimelineEventType.OTHER: "comment",
        }
        return mapping.get(event_type, "comment")

    @with_session
    def add_timeline_event(
        self,
        incident_id: int,
        event_type: TimelineEventType,
        description: str,
        operator: str,
        session: Session = None
    ) -> IncidentTimeline:
        self.logger.info(
            f"Adding timeline event to incident {incident_id}: {event_type.value}"
        )

        timeline_event = IncidentTimeline(
            incident_id=incident_id,
            event_type=self._map_event_type(event_type),
            description=description,
            operator=operator
        )
        session.add(timeline_event)
        session.flush()

        log_audit(
            action="add_timeline_event",
            resource_type="incident",
            resource_id=str(incident_id),
            detail=f"Event: {event_type.value}, Description: {description[:100]}",
            user=operator
        )

        return timeline_event

    @with_read_session
    def get_timeline(
        self,
        incident_id: int,
        session: Session = None
    ) -> List[TimelineEvent]:
        timelines = session.query(IncidentTimeline).filter(
            IncidentTimeline.incident_id == incident_id
        ).order_by(IncidentTimeline.created_at).all()

        event_type_mapping = {
            "detected": TimelineEventType.DETECTED,
            "triaged": TimelineEventType.OTHER,
            "escalated": TimelineEventType.NOTIFIED,
            "mitigated": TimelineEventType.CONTAINED,
            "resolved": TimelineEventType.RECOVERED,
            "comment": TimelineEventType.COMMUNICATED,
        }

        return [
            TimelineEvent(
                event_type=event_type_mapping.get(t.event_type, TimelineEventType.OTHER),
                description=t.description,
                operator=t.operator,
                timestamp=t.created_at
            )
            for t in timelines
        ]

    def auto_record_milestone(
        self,
        incident_id: int,
        status: IncidentStatus,
        operator: str,
        remark: str = "",
        session: Session = None
    ) -> None:
        status_event_map = {
            IncidentStatus.OPEN: (TimelineEventType.DETECTED, f"事件创建并打开: {remark}"),
            IncidentStatus.INVESTIGATING: (TimelineEventType.DETECTED, f"开始调查: {remark}"),
            IncidentStatus.MITIGATED: (TimelineEventType.CONTAINED, f"已遏制: {remark}"),
            IncidentStatus.RESOLVED: (TimelineEventType.ERADICATED, f"已根除: {remark}"),
            IncidentStatus.CLOSED: (TimelineEventType.RECOVERED, f"事件关闭: {remark}"),
        }

        if status in status_event_map:
            event_type, description = status_event_map[status]
            self.add_timeline_event(
                incident_id=incident_id,
                event_type=event_type,
                description=description if remark else description.replace(": ", ""),
                operator=operator,
                session=session
            )


class IncidentManager:
    def __init__(self):
        self.logger = logger
        self.timeline_manager = TimelineManager()

    @staticmethod
    def _map_incident_type(incident_type: IncidentType) -> str:
        mapping = {
            IncidentType.DATA_BREACH: "data_breach",
            IncidentType.INTRUSION: "unauthorized_access",
            IncidentType.RANSOMWARE: "malware",
            IncidentType.DDOS: "dos_attack",
            IncidentType.COMPLIANCE_VIOLATION: "other",
            IncidentType.OTHER: "other",
        }
        return mapping.get(incident_type, "other")

    @staticmethod
    def _map_incident_status(status: IncidentStatus) -> str:
        mapping = {
            IncidentStatus.OPEN: "open",
            IncidentStatus.INVESTIGATING: "investigating",
            IncidentStatus.MITIGATED: "contained",
            IncidentStatus.RESOLVED: "recovered",
            IncidentStatus.CLOSED: "closed",
        }
        return mapping.get(status, "open")

    @with_session
    @with_log_context(operation_type="create_incident")
    def create_incident(
        self,
        title: str,
        description: str,
        incident_type: IncidentType,
        severity: SeverityEnum,
        assets_affected: List[int],
        created_by: str,
        assigned_to: Optional[str] = None,
        session: Session = None
    ) -> Incident:
        self.logger.info(f"Creating incident: {title}")

        assets_str = ",".join(map(str, assets_affected)) if assets_affected else None

        incident = Incident(
            title=title,
            description=description,
            type=self._map_incident_type(incident_type),
            severity=severity,
            status=self._map_incident_status(IncidentStatus.OPEN),
            assets_affected=assets_str,
            created_by=created_by,
            assigned_to=assigned_to
        )
        session.add(incident)
        session.flush()

        self.timeline_manager.add_timeline_event(
            incident_id=incident.id,
            event_type=TimelineEventType.DETECTED,
            description=f"事件由 {created_by} 创建",
            operator=created_by,
            session=session
        )

        if assigned_to:
            self.timeline_manager.add_timeline_event(
                incident_id=incident.id,
                event_type=TimelineEventType.NOTIFIED,
                description=f"事件指派给 {assigned_to}",
                operator=created_by,
                session=session
            )

        log_audit(
            action="create_incident",
            resource_type="incident",
            resource_id=str(incident.id),
            detail=f"Title: {title}, Type: {incident_type.value}, Severity: {severity.value}",
            user=created_by
        )

        return incident

    @with_session
    @with_log_context(operation_type="update_incident_status")
    def update_incident_status(
        self,
        incident_id: int,
        new_status: IncidentStatus,
        operator: str,
        remark: str = "",
        session: Session = None
    ) -> Incident:
        self.logger.info(f"Updating incident {incident_id} status to {new_status.value}")

        incident = session.query(Incident).filter(Incident.id == incident_id).first()
        if not incident:
            raise ValueError(f"Incident {incident_id} not found")

        incident.status = self._map_incident_status(new_status)
        incident.updated_at = datetime.now(timezone.utc)

        if new_status == IncidentStatus.RESOLVED:
            incident.resolved_at = datetime.now(timezone.utc)
        elif new_status == IncidentStatus.CLOSED:
            incident.closed_at = datetime.now(timezone.utc)

        self.timeline_manager.auto_record_milestone(
            incident_id=incident_id,
            status=new_status,
            operator=operator,
            remark=remark,
            session=session
        )

        log_audit(
            action="update_incident_status",
            resource_type="incident",
            resource_id=str(incident_id),
            detail=f"New status: {new_status.value}, Remark: {remark}",
            user=operator
        )

        return incident

    @with_session
    def assign_incident(
        self,
        incident_id: int,
        assigned_to: str,
        operator: str,
        session: Session = None
    ) -> Incident:
        self.logger.info(f"Assigning incident {incident_id} to {assigned_to}")

        incident = session.query(Incident).filter(Incident.id == incident_id).first()
        if not incident:
            raise ValueError(f"Incident {incident_id} not found")

        incident.assigned_to = assigned_to
        incident.updated_at = datetime.now(timezone.utc)

        self.timeline_manager.add_timeline_event(
            incident_id=incident_id,
            event_type=TimelineEventType.NOTIFIED,
            description=f"事件重新指派给 {assigned_to}",
            operator=operator,
            session=session
        )

        log_audit(
            action="assign_incident",
            resource_type="incident",
            resource_id=str(incident_id),
            detail=f"Assigned to: {assigned_to}",
            user=operator
        )

        return incident

    @with_read_session
    def get_incident(
        self,
        incident_id: int,
        session: Session = None
    ) -> Optional[Incident]:
        return session.query(Incident).filter(Incident.id == incident_id).first()

    @with_session
    def update_incident(
        self,
        incident_id: int,
        updates: Dict[str, Any],
        operator: str,
        session: Session = None
    ) -> Incident:
        incident = session.query(Incident).filter(Incident.id == incident_id).first()
        if not incident:
            raise ValueError(f"Incident {incident_id} not found")

        for key, value in updates.items():
            if key == "incident_type":
                setattr(incident, "type", self._map_incident_type(value))
            elif key == "status":
                setattr(incident, "status", self._map_incident_status(value))
            elif hasattr(incident, key) and key not in ["id", "created_at"]:
                setattr(incident, key, value)

        incident.updated_at = datetime.now(timezone.utc)

        log_audit(
            action="update_incident",
            resource_type="incident",
            resource_id=str(incident_id),
            detail=f"Updates: {json.dumps(updates, default=str, ensure_ascii=False)}",
            user=operator
        )

        return incident


class IncidentAnalyzer:
    def __init__(self):
        self.logger = logger
        self.timeline_manager = TimelineManager()

    @with_read_session
    def get_affected_assets_details(
        self,
        incident: Incident,
        session: Session = None
    ) -> List[Dict[str, Any]]:
        if not incident.assets_affected:
            return []

        asset_ids = [int(a) for a in incident.assets_affected.split(",") if a.isdigit()]
        assets = session.query(Asset).filter(Asset.id.in_(asset_ids)).all()

        return [
            {
                "id": asset.id,
                "name": asset.name,
                "ip": asset.ip,
                "type": asset.type,
                "importance": asset.importance,
                "owner": asset.owner,
                "department": asset.department,
                "description": asset.description
            }
            for asset in assets
        ]

    @with_read_session
    def get_related_vulnerabilities(
        self,
        incident: Incident,
        days: int = 90,
        session: Session = None
    ) -> List[Dict[str, Any]]:
        if not incident.assets_affected:
            return []

        asset_ids = [int(a) for a in incident.assets_affected.split(",") if a.isdigit()]
        cutoff_date = datetime.now(timezone.utc) - timedelta(days=days)

        vuln_instances = session.query(VulnerabilityInstance).filter(
            and_(
                VulnerabilityInstance.asset_id.in_(asset_ids),
                VulnerabilityInstance.discovery_time >= cutoff_date
            )
        ).all()

        related_vulns = []
        for vi in vuln_instances:
            vuln = vi.vulnerability
            asset = vi.asset
            related_vulns.append({
                "vuln_instance_id": vi.id,
                "vuln_id": vuln.id,
                "cve_id": vuln.cve_id,
                "title": vuln.title,
                "severity": vuln.severity.value,
                "cvss_score": float(vi.risk_score),
                "asset_name": asset.name,
                "asset_ip": asset.ip,
                "discovery_time": vi.discovery_time.isoformat(),
                "fix_status": vi.fix_status.value
            })

        return related_vulns

    def _calculate_metrics(
        self,
        timeline: List[TimelineEvent]
    ) -> Tuple[Optional[float], Optional[float]]:
        detected_time = None
        first_response_time = None
        resolved_time = None

        for event in timeline:
            if event.event_type == TimelineEventType.DETECTED and detected_time is None:
                detected_time = event.timestamp
            elif event.event_type in [TimelineEventType.NOTIFIED, TimelineEventType.DETECTED] and first_response_time is None:
                if detected_time and event.timestamp > detected_time:
                    first_response_time = event.timestamp
            elif event.event_type == TimelineEventType.RECOVERED:
                resolved_time = event.timestamp

        mtta = None
        mttr = None

        if detected_time and first_response_time:
            mtta = (first_response_time - detected_time).total_seconds() / 3600

        if detected_time and resolved_time:
            mttr = (resolved_time - detected_time).total_seconds() / 3600

        return mtta, mttr

    def _generate_timeline_milestones(
        self,
        timeline: List[TimelineEvent]
    ) -> Dict[str, Optional[datetime]]:
        milestones = {
            "detected_time": None,
            "first_response_time": None,
            "contained_time": None,
            "eradicated_time": None,
            "recovered_time": None,
            "closed_time": None,
        }

        for event in timeline:
            if event.event_type == TimelineEventType.DETECTED and milestones["detected_time"] is None:
                milestones["detected_time"] = event.timestamp
            elif event.event_type == TimelineEventType.NOTIFIED and milestones["first_response_time"] is None:
                milestones["first_response_time"] = event.timestamp
            elif event.event_type == TimelineEventType.CONTAINED and milestones["contained_time"] is None:
                milestones["contained_time"] = event.timestamp
            elif event.event_type == TimelineEventType.ERADICATED and milestones["eradicated_time"] is None:
                milestones["eradicated_time"] = event.timestamp
            elif event.event_type == TimelineEventType.RECOVERED and milestones["recovered_time"] is None:
                milestones["recovered_time"] = event.timestamp

        return milestones

    @with_read_session
    @with_log_context(operation_type="analyze_incident")
    def analyze_incident(
        self,
        incident_id: int,
        session: Session = None
    ) -> IncidentReport:
        self.logger.info(f"Analyzing incident {incident_id}")

        incident = session.query(Incident).filter(Incident.id == incident_id).first()
        if not incident:
            raise ValueError(f"Incident {incident_id} not found")

        timeline = self.timeline_manager.get_timeline(incident_id=incident_id, session=session)
        affected_assets = self.get_affected_assets_details(incident=incident, session=session)
        related_vulns = self.get_related_vulnerabilities(incident=incident, session=session)

        mtta, mttr = self._calculate_metrics(timeline)
        milestones = self._generate_timeline_milestones(timeline)

        affected_users_count = len(set(asset["owner"] for asset in affected_assets))

        response_measures = []
        for event in timeline:
            if event.event_type in [TimelineEventType.CONTAINED, TimelineEventType.ERADICATED, TimelineEventType.RECOVERED]:
                response_measures.append(event.description)

        status_map = {
            "open": IncidentStatus.OPEN,
            "investigating": IncidentStatus.INVESTIGATING,
            "contained": IncidentStatus.MITIGATED,
            "eradicated": IncidentStatus.RESOLVED,
            "recovered": IncidentStatus.RESOLVED,
            "closed": IncidentStatus.CLOSED,
        }
        current_status = status_map.get(incident.status, IncidentStatus.OPEN)

        type_map = {
            "data_breach": IncidentType.DATA_BREACH,
            "malware": IncidentType.RANSOMWARE,
            "unauthorized_access": IncidentType.INTRUSION,
            "dos_attack": IncidentType.DDOS,
            "phishing": IncidentType.OTHER,
            "insider_threat": IncidentType.INTRUSION,
            "other": IncidentType.OTHER,
        }
        current_type = type_map.get(incident.type, IncidentType.OTHER)

        report = IncidentReport(
            incident_id=incident.id,
            title=incident.title,
            incident_type=current_type,
            severity=incident.severity,
            status=current_status,
            timeline=timeline,
            affected_assets=affected_assets,
            affected_users_count=affected_users_count,
            related_vulnerabilities=related_vulns,
            response_measures=response_measures,
            root_cause=self._generate_root_cause_analysis(related_vulns, timeline),
            improvement_suggestions=self._generate_improvement_suggestions(related_vulns, affected_assets),
            mtta=mtta,
            mttr=mttr,
            business_impact=self._generate_business_impact(affected_assets, related_vulns)
        )

        log_audit(
            action="analyze_incident",
            resource_type="incident",
            resource_id=str(incident_id),
            detail=f"Generated analysis report",
            user="system"
        )

        return report

    def _generate_root_cause_analysis(
        self,
        related_vulns: List[Dict[str, Any]],
        timeline: List[TimelineEvent]
    ) -> str:
        if not related_vulns:
            return "未发现相关漏洞，建议进一步调查入侵来源。"

        high_risk_vulns = [v for v in related_vulns if v["severity"] in ["critical", "high"]]
        if high_risk_vulns:
            vuln_titles = "; ".join(v["title"] for v in high_risk_vulns[:3])
            return f"可能由以下高危漏洞导致: {vuln_titles}。建议进行全面的漏洞扫描和渗透测试。"

        return "初步分析显示可能与已知漏洞相关，建议进行深度取证分析以确定确切原因。"

    def _generate_improvement_suggestions(
        self,
        related_vulns: List[Dict[str, Any]],
        affected_assets: List[Dict[str, Any]]
    ) -> List[str]:
        suggestions = [
            "加强边界防护，部署入侵检测/防御系统(IDS/IPS)",
            "定期进行安全意识培训，防范社会工程学攻击",
            "实施多因素认证，增强账户安全",
            "建立完善的备份策略，确保数据可恢复",
            "制定并定期演练应急响应预案",
        ]

        if related_vulns:
            suggestions.insert(0, "及时修复所有相关漏洞，特别是高危漏洞")
            suggestions.insert(1, "建立定期漏洞扫描机制，实现漏洞闭环管理")

        high_importance_assets = [a for a in affected_assets if a.get("importance", 0) >= 4]
        if high_importance_assets:
            suggestions.append("对重要资产实施额外的安全防护措施和监控")

        return suggestions

    def _generate_business_impact(
        self,
        affected_assets: List[Dict[str, Any]],
        related_vulns: List[Dict[str, Any]]
    ) -> str:
        if not affected_assets:
            return "暂无受影响资产信息"

        departments = set(a["department"] for a in affected_assets)
        high_importance = [a for a in affected_assets if a.get("importance", 0) >= 4]

        impact_parts = [
            f"受影响部门: {', '.join(departments)}",
            f"受影响资产数量: {len(affected_assets)}",
        ]

        if high_importance:
            impact_parts.append(f"重要资产受影响: {len(high_importance)} 个")

        if related_vulns:
            critical_count = sum(1 for v in related_vulns if v["severity"] == "critical")
            high_count = sum(1 for v in related_vulns if v["severity"] == "high")
            impact_parts.append(f"相关高危漏洞: {critical_count} 个严重, {high_count} 个高危")

        return "; ".join(impact_parts)


class QueryEngine:
    def __init__(self):
        self.logger = logger

    def _apply_query_filters(
        self,
        query,
        model,
        filters: QueryFilter,
        join_models: Optional[Dict[str, Any]] = None
    ):
        conditions = []

        if filters.vuln_id is not None:
            if hasattr(model, "vuln_id"):
                conditions.append(model.vuln_id == filters.vuln_id)
            elif hasattr(model, "id") and join_models and "vulnerability" in join_models:
                conditions.append(join_models["vulnerability"].id == filters.vuln_id)

        if filters.cve_id:
            if hasattr(model, "cve_id"):
                conditions.append(model.cve_id.like(f"%{filters.cve_id}%"))
            elif join_models and "vulnerability" in join_models:
                conditions.append(join_models["vulnerability"].cve_id.like(f"%{filters.cve_id}%"))

        if filters.asset_name:
            if join_models and "asset" in join_models:
                conditions.append(join_models["asset"].name.like(f"%{filters.asset_name}%"))

        if filters.asset_ip:
            if join_models and "asset" in join_models:
                conditions.append(join_models["asset"].ip.like(f"%{filters.asset_ip}%"))

        if filters.department:
            if join_models and "asset" in join_models:
                conditions.append(join_models["asset"].department.like(f"%{filters.department}%"))

        if filters.importance is not None:
            if join_models and "asset" in join_models:
                conditions.append(join_models["asset"].importance == filters.importance)

        if filters.discovery_time_start:
            if hasattr(model, "discovery_time"):
                conditions.append(model.discovery_time >= filters.discovery_time_start)
            elif hasattr(model, "created_at"):
                conditions.append(model.created_at >= filters.discovery_time_start)

        if filters.discovery_time_end:
            if hasattr(model, "discovery_time"):
                conditions.append(model.discovery_time <= filters.discovery_time_end)
            elif hasattr(model, "created_at"):
                conditions.append(model.created_at <= filters.discovery_time_end)

        if filters.fix_time_start and hasattr(model, "fixed_at"):
            conditions.append(model.fixed_at >= filters.fix_time_start)

        if filters.fix_time_end and hasattr(model, "fixed_at"):
            conditions.append(model.fixed_at <= filters.fix_time_end)

        if filters.close_time_start and hasattr(model, "closed_at"):
            conditions.append(model.closed_at >= filters.close_time_start)

        if filters.close_time_end and hasattr(model, "closed_at"):
            conditions.append(model.closed_at <= filters.close_time_end)

        if filters.work_order_status:
            if hasattr(model, "status") and isinstance(model.status.type, type(WorkOrderStatusEnum.PENDING)):
                conditions.append(model.status == filters.work_order_status)

        if filters.risk_level:
            if hasattr(model, "severity"):
                conditions.append(model.severity == filters.risk_level)
            elif join_models and "vulnerability" in join_models:
                conditions.append(join_models["vulnerability"].severity == filters.risk_level)

        if filters.severity:
            if hasattr(model, "severity"):
                conditions.append(model.severity == filters.severity)
            elif join_models and "vulnerability" in join_models:
                conditions.append(join_models["vulnerability"].severity == filters.severity)

        if filters.operator:
            if hasattr(model, "operator"):
                conditions.append(model.operator.like(f"%{filters.operator}%"))
            elif hasattr(model, "created_by"):
                conditions.append(model.created_by.like(f"%{filters.operator}%"))

        if filters.assignee:
            if hasattr(model, "assignee"):
                conditions.append(model.assignee.like(f"%{filters.assignee}%"))
            elif hasattr(model, "assigned_to"):
                conditions.append(model.assigned_to.like(f"%{filters.assignee}%"))

        if filters.custom_filters:
            for key, value in filters.custom_filters.items():
                if hasattr(model, key):
                    if isinstance(value, str):
                        conditions.append(getattr(model, key).like(f"%{value}%"))
                    else:
                        conditions.append(getattr(model, key) == value)

        if conditions:
            query = query.filter(and_(*conditions))

        return query

    @with_read_session
    def query_vulnerabilities(
        self,
        filters: QueryFilter,
        page: int = 1,
        page_size: int = 20,
        order_by: str = "created_at",
        order_dir: str = "desc",
        fields: Optional[List[str]] = None,
        session: Session = None
    ) -> Dict[str, Any]:
        self.logger.info(f"Querying vulnerabilities with filters: {asdict(filters)}")

        query = session.query(
            VulnerabilityInstance,
            Vulnerability,
            Asset
        ).join(
            Vulnerability, VulnerabilityInstance.vuln_id == Vulnerability.id
        ).join(
            Asset, VulnerabilityInstance.asset_id == Asset.id
        )

        join_models = {"vulnerability": Vulnerability, "asset": Asset}
        query = self._apply_query_filters(query, VulnerabilityInstance, filters, join_models)

        total = query.count()

        order_func = desc if order_dir.lower() == "desc" else asc
        if hasattr(VulnerabilityInstance, order_by):
            query = query.order_by(order_func(getattr(VulnerabilityInstance, order_by)))
        elif hasattr(Vulnerability, order_by):
            query = query.order_by(order_func(getattr(Vulnerability, order_by)))
        elif hasattr(Asset, order_by):
            query = query.order_by(order_func(getattr(Asset, order_by)))
        else:
            query = query.order_by(order_func(VulnerabilityInstance.created_at))

        offset = (page - 1) * page_size
        query = query.offset(offset).limit(page_size)

        results = query.all()

        items = []
        for vi, vuln, asset in results:
            item = {
                "vuln_instance_id": vi.id,
                "vuln_id": vuln.id,
                "cve_id": vuln.cve_id,
                "title": vuln.title,
                "description": vuln.description,
                "severity": vuln.severity.value,
                "cvss_score": float(vuln.cvss_score) if vuln.cvss_score else None,
                "risk_score": float(vi.risk_score),
                "asset_id": asset.id,
                "asset_name": asset.name,
                "asset_ip": asset.ip,
                "asset_type": asset.type,
                "department": asset.department,
                "importance": asset.importance,
                "discovery_time": vi.discovery_time.isoformat(),
                "fix_deadline": vi.fix_deadline.isoformat(),
                "fix_status": vi.fix_status.value,
                "port": vi.port,
                "protocol": vi.protocol,
                "location": vi.location,
            }
            if fields:
                item = {k: v for k, v in item.items() if k in fields}
            items.append(item)

        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size,
            "items": items
        }

    @with_read_session
    def query_work_orders(
        self,
        filters: QueryFilter,
        page: int = 1,
        page_size: int = 20,
        order_by: str = "created_at",
        order_dir: str = "desc",
        fields: Optional[List[str]] = None,
        session: Session = None
    ) -> Dict[str, Any]:
        self.logger.info(f"Querying work orders with filters: {asdict(filters)}")

        query = session.query(
            WorkOrder,
            VulnerabilityInstance,
            Vulnerability,
            Asset
        ).join(
            VulnerabilityInstance, WorkOrder.vuln_instance_id == VulnerabilityInstance.id
        ).join(
            Vulnerability, VulnerabilityInstance.vuln_id == Vulnerability.id
        ).join(
            Asset, VulnerabilityInstance.asset_id == Asset.id
        )

        join_models = {"vulnerability": Vulnerability, "asset": Asset}
        query = self._apply_query_filters(query, WorkOrder, filters, join_models)

        total = query.count()

        order_func = desc if order_dir.lower() == "desc" else asc
        if hasattr(WorkOrder, order_by):
            query = query.order_by(order_func(getattr(WorkOrder, order_by)))
        elif hasattr(VulnerabilityInstance, order_by):
            query = query.order_by(order_func(getattr(VulnerabilityInstance, order_by)))
        else:
            query = query.order_by(order_func(WorkOrder.created_at))

        offset = (page - 1) * page_size
        query = query.offset(offset).limit(page_size)

        results = query.all()

        items = []
        for wo, vi, vuln, asset in results:
            item = {
                "work_order_id": wo.id,
                "vuln_instance_id": vi.id,
                "vuln_title": vuln.title,
                "cve_id": vuln.cve_id,
                "severity": vuln.severity.value,
                "asset_name": asset.name,
                "asset_ip": asset.ip,
                "department": asset.department,
                "assignee": wo.assignee,
                "status": wo.status.value,
                "priority": wo.priority,
                "escalation_level": wo.escalation_level,
                "created_at": wo.created_at.isoformat(),
                "started_at": wo.started_at.isoformat() if wo.started_at else None,
                "fixed_at": wo.fixed_at.isoformat() if wo.fixed_at else None,
                "verified_at": wo.verified_at.isoformat() if wo.verified_at else None,
                "closed_at": wo.closed_at.isoformat() if wo.closed_at else None,
                "deadline": wo.deadline.isoformat(),
                "remarks": wo.remarks,
            }
            if fields:
                item = {k: v for k, v in item.items() if k in fields}
            items.append(item)

        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size,
            "items": items
        }

    @with_read_session
    def query_incidents(
        self,
        filters: QueryFilter,
        page: int = 1,
        page_size: int = 20,
        order_by: str = "created_at",
        order_dir: str = "desc",
        fields: Optional[List[str]] = None,
        session: Session = None
    ) -> Dict[str, Any]:
        self.logger.info(f"Querying incidents with filters: {asdict(filters)}")

        query = session.query(Incident)

        if filters.incident_type:
            type_mapping = {
                IncidentType.DATA_BREACH: "data_breach",
                IncidentType.INTRUSION: "unauthorized_access",
                IncidentType.RANSOMWARE: "malware",
                IncidentType.DDOS: "dos_attack",
                IncidentType.COMPLIANCE_VIOLATION: "other",
                IncidentType.OTHER: "other",
            }
            query = query.filter(Incident.type == type_mapping[filters.incident_type])

        if filters.incident_status:
            status_mapping = {
                IncidentStatus.OPEN: "open",
                IncidentStatus.INVESTIGATING: "investigating",
                IncidentStatus.MITIGATED: "contained",
                IncidentStatus.RESOLVED: "recovered",
                IncidentStatus.CLOSED: "closed",
            }
            query = query.filter(Incident.status == status_mapping[filters.incident_status])

        query = self._apply_query_filters(query, Incident, filters)

        total = query.count()

        order_func = desc if order_dir.lower() == "desc" else asc
        if hasattr(Incident, order_by):
            query = query.order_by(order_func(getattr(Incident, order_by)))
        else:
            query = query.order_by(order_func(Incident.created_at))

        offset = (page - 1) * page_size
        query = query.offset(offset).limit(page_size)

        results = query.all()

        items = []
        for incident in results:
            asset_ids = []
            if incident.assets_affected:
                asset_ids = [int(a) for a in incident.assets_affected.split(",") if a.isdigit()]

            asset_names = []
            if asset_ids:
                assets = session.query(Asset).filter(Asset.id.in_(asset_ids)).all()
                asset_names = [a.name for a in assets]

            item = {
                "incident_id": incident.id,
                "title": incident.title,
                "description": incident.description,
                "type": incident.type,
                "severity": incident.severity.value,
                "status": incident.status,
                "assets_affected": asset_ids,
                "asset_names": asset_names,
                "created_by": incident.created_by,
                "assigned_to": incident.assigned_to,
                "created_at": incident.created_at.isoformat(),
                "updated_at": incident.updated_at.isoformat(),
                "resolved_at": incident.resolved_at.isoformat() if incident.resolved_at else None,
                "closed_at": incident.closed_at.isoformat() if incident.closed_at else None,
            }
            if fields:
                item = {k: v for k, v in item.items() if k in fields}
            items.append(item)

        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size,
            "items": items
        }

    @with_read_session
    def query_lifecycle_records(
        self,
        vuln_instance_id: int,
        session: Session = None
    ) -> Dict[str, Any]:
        self.logger.info(f"Querying lifecycle records for vuln_instance {vuln_instance_id}")

        vuln_instance = session.query(VulnerabilityInstance).filter(
            VulnerabilityInstance.id == vuln_instance_id
        ).first()

        if not vuln_instance:
            raise ValueError(f"Vulnerability instance {vuln_instance_id} not found")

        vuln = vuln_instance.vulnerability
        asset = vuln_instance.asset

        work_orders = session.query(WorkOrder).filter(
            WorkOrder.vuln_instance_id == vuln_instance_id
        ).order_by(WorkOrder.created_at).all()

        notifications = []
        escalation_records = []
        for wo in work_orders:
            notifications.extend(wo.notifications)
            escalation_records.extend(wo.escalation_records)

        verification_records = session.query(VerificationRecord).filter(
            VerificationRecord.vuln_instance_id == vuln_instance_id
        ).order_by(VerificationRecord.verification_time).all()

        review_tasks = session.query(ReviewTask).filter(
            ReviewTask.vuln_instance_id == vuln_instance_id
        ).order_by(ReviewTask.created_at).all()

        response_plans = session.query(ResponsePlan).filter(
            ResponsePlan.vuln_instance_id == vuln_instance_id
        ).order_by(ResponsePlan.created_at).all()

        audit_logs = session.query(AuditLog).filter(
            and_(
                AuditLog.resource_type == "vulnerability_instance",
                AuditLog.resource_id == str(vuln_instance_id)
            )
        ).order_by(AuditLog.created_at).all()

        status_history = self._build_status_history(work_orders, verification_records)

        return {
            "vulnerability_info": {
                "vuln_id": vuln.id,
                "cve_id": vuln.cve_id,
                "title": vuln.title,
                "description": vuln.description,
                "severity": vuln.severity.value,
                "cvss_score": float(vuln.cvss_score) if vuln.cvss_score else None,
                "cwe_id": vuln.cwe_id,
                "source": vuln.source,
                "first_seen": vuln.first_seen.isoformat(),
                "last_seen": vuln.last_seen.isoformat(),
            },
            "asset_info": {
                "asset_id": asset.id,
                "name": asset.name,
                "ip": asset.ip,
                "type": asset.type,
                "importance": asset.importance,
                "owner": asset.owner,
                "department": asset.department,
            },
            "risk_assessment": {
                "risk_score": float(vuln_instance.risk_score),
                "discovery_time": vuln_instance.discovery_time.isoformat(),
                "fix_deadline": vuln_instance.fix_deadline.isoformat(),
                "fix_status": vuln_instance.fix_status.value,
                "port": vuln_instance.port,
                "protocol": vuln_instance.protocol,
                "location": vuln_instance.location,
            },
            "work_orders": [
                {
                    "id": wo.id,
                    "assignee": wo.assignee,
                    "status": wo.status.value,
                    "priority": wo.priority,
                    "escalation_level": wo.escalation_level,
                    "created_at": wo.created_at.isoformat(),
                    "started_at": wo.started_at.isoformat() if wo.started_at else None,
                    "fixed_at": wo.fixed_at.isoformat() if wo.fixed_at else None,
                    "verified_at": wo.verified_at.isoformat() if wo.verified_at else None,
                    "closed_at": wo.closed_at.isoformat() if wo.closed_at else None,
                    "deadline": wo.deadline.isoformat(),
                }
                for wo in work_orders
            ],
            "status_history": status_history,
            "notifications": [
                {
                    "id": n.id,
                    "type": n.type.value,
                    "recipient": n.recipient,
                    "status": n.status.value,
                    "sent_at": n.sent_at.isoformat() if n.sent_at else None,
                    "escalation_level": n.escalation_level,
                    "created_at": n.created_at.isoformat(),
                }
                for n in notifications
            ],
            "verification_records": [
                {
                    "id": vr.id,
                    "work_order_id": vr.work_order_id,
                    "scan_type": vr.scan_type,
                    "is_fixed": vr.is_fixed,
                    "details": vr.details,
                    "operator": vr.operator,
                    "verification_time": vr.verification_time.isoformat(),
                }
                for vr in verification_records
            ],
            "escalation_records": [
                {
                    "id": er.id,
                    "work_order_id": er.work_order_id,
                    "old_level": er.old_level,
                    "new_level": er.new_level,
                    "reason": er.reason,
                    "escalated_by": er.escalated_by,
                    "created_at": er.created_at.isoformat(),
                }
                for er in escalation_records
            ],
            "review_tasks": [
                {
                    "id": rt.id,
                    "work_order_id": rt.work_order_id,
                    "reason": rt.reason.value,
                    "status": rt.status.value,
                    "assignees": rt.assignees,
                    "root_cause": rt.root_cause,
                    "improvement_measures": rt.improvement_measures,
                    "deadline": rt.deadline.isoformat(),
                    "completed_at": rt.completed_at.isoformat() if rt.completed_at else None,
                    "created_by": rt.created_by,
                    "created_at": rt.created_at.isoformat(),
                }
                for rt in review_tasks
            ],
            "response_plans": [
                {
                    "id": rp.id,
                    "incident_id": rp.incident_id,
                    "status": rp.status.value,
                    "trigger_reason": rp.trigger_reason,
                    "isolation_measures": rp.isolation_measures,
                    "mitigation_measures": rp.mitigation_measures,
                    "root_fix_plan": rp.root_fix_plan,
                    "execution_time": rp.execution_time.isoformat(),
                    "effectiveness": rp.effectiveness,
                    "completed_at": rp.completed_at.isoformat() if rp.completed_at else None,
                }
                for rp in response_plans
            ],
            "audit_logs": [
                {
                    "id": al.id,
                    "user": al.user,
                    "action": al.action,
                    "detail": al.detail,
                    "ip": al.ip,
                    "created_at": al.created_at.isoformat(),
                }
                for al in audit_logs
            ],
        }

    def _build_status_history(
        self,
        work_orders: List[WorkOrder],
        verification_records: List[VerificationRecord]
    ) -> List[Dict[str, Any]]:
        history = []

        for wo in work_orders:
            transitions = [
                (wo.created_at, "created", f"工单创建，分配给 {wo.assignee}"),
                (wo.started_at, "started", "开始修复"),
                (wo.fixed_at, "fixed", "修复完成"),
                (wo.verified_at, "verified", "验证完成"),
                (wo.closed_at, "closed", "工单关闭"),
            ]
            for time, status, description in transitions:
                if time:
                    history.append({
                        "timestamp": time.isoformat(),
                        "status": status,
                        "description": description,
                        "work_order_id": wo.id
                    })

        for vr in verification_records:
            history.append({
                "timestamp": vr.verification_time.isoformat(),
                "status": "verification",
                "description": f"验证结果: {'已修复' if vr.is_fixed else '未修复'} - {vr.details or ''}",
                "work_order_id": vr.work_order_id
            })

        history.sort(key=lambda x: x["timestamp"])
        return history

    @with_read_session
    def query_audit_logs(
        self,
        filters: QueryFilter,
        page: int = 1,
        page_size: int = 20,
        order_by: str = "created_at",
        order_dir: str = "desc",
        fields: Optional[List[str]] = None,
        session: Session = None
    ) -> Dict[str, Any]:
        query = session.query(AuditLog)
        query = self._apply_query_filters(query, AuditLog, filters)

        total = query.count()

        order_func = desc if order_dir.lower() == "desc" else asc
        query = query.order_by(order_func(AuditLog.created_at))

        offset = (page - 1) * page_size
        query = query.offset(offset).limit(page_size)

        results = query.all()

        items = []
        for log in results:
            item = {
                "id": log.id,
                "user": log.user,
                "action": log.action,
                "resource_type": log.resource_type,
                "resource_id": log.resource_id,
                "detail": log.detail,
                "ip": log.ip,
                "user_agent": log.user_agent,
                "request_id": log.request_id,
                "created_at": log.created_at.isoformat(),
            }
            if fields:
                item = {k: v for k, v in item.items() if k in fields}
            items.append(item)

        return {
            "total": total,
            "page": page,
            "page_size": page_size,
            "total_pages": (total + page_size - 1) // page_size,
            "items": items
        }


class ExportManager:
    def __init__(self):
        self.logger = logger
        self.query_engine = QueryEngine()
        self._export_tasks: Dict[str, ExportTask] = {}
        self._ensure_export_dirs()

    def _ensure_export_dirs(self):
        os.makedirs(config.REPORT_OUTPUT_DIR, exist_ok=True)
        os.makedirs(config.EXPORT_TEMP_DIR, exist_ok=True)

    def _generate_task_id(self) -> str:
        return f"export_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}_{os.urandom(4).hex()}"

    def _get_query_method(self, query_type: QueryType):
        method_map = {
            QueryType.VULNERABILITIES: self.query_engine.query_vulnerabilities,
            QueryType.WORK_ORDERS: self.query_engine.query_work_orders,
            QueryType.INCIDENTS: self.query_engine.query_incidents,
        }
        return method_map.get(query_type)

    def _get_all_data(
        self,
        query_type: QueryType,
        filters: QueryFilter,
        fields: Optional[List[str]] = None,
        batch_size: int = 1000
    ) -> List[Dict[str, Any]]:
        query_method = self._get_query_method(query_type)
        if not query_method:
            raise ValueError(f"Unsupported query type: {query_type}")

        all_data = []
        page = 1

        while True:
            result = query_method(
                filters=filters,
                page=page,
                page_size=batch_size,
                fields=fields
            )
            all_data.extend(result["items"])

            if page >= result["total_pages"]:
                break
            page += 1

        return all_data

    def _get_lifecycle_data(
        self,
        vuln_instance_ids: List[int],
        fields: Optional[List[str]] = None
    ) -> List[Dict[str, Any]]:
        all_data = []
        for vid in vuln_instance_ids:
            try:
                record = self.query_engine.query_lifecycle_records(vid)
                if fields:
                    record = {k: v for k, v in record.items() if k in fields}
                all_data.append(record)
            except Exception as e:
                self.logger.error(f"Failed to get lifecycle record for {vid}: {e}")
        return all_data

    def _export_to_csv(
        self,
        data: List[Dict[str, Any]],
        file_path: str,
        fields: Optional[List[str]] = None
    ) -> None:
        if not data:
            with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(fields or [])
            return

        if fields:
            actual_fields = [f for f in fields if f in data[0].keys()]
        else:
            actual_fields = list(data[0].keys())

        with open(file_path, "w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=actual_fields)
            writer.writeheader()
            for row in data:
                filtered_row = {k: row.get(k, "") for k in actual_fields}
                writer.writerow(filtered_row)

    def _export_to_excel(
        self,
        data: List[Dict[str, Any]],
        file_path: str,
        sheet_name: str = "Data",
        fields: Optional[List[str]] = None
    ) -> None:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name

        if not data:
            if fields:
                ws.append(fields)
            wb.save(file_path)
            return

        if fields:
            actual_fields = [f for f in fields if f in data[0].keys()]
        else:
            actual_fields = list(data[0].keys())

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        ws.append(actual_fields)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        for row in data:
            filtered_row = [row.get(k, "") for k in actual_fields]
            ws.append(filtered_row)

        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        wb.save(file_path)

    def _export_to_json(
        self,
        data: List[Dict[str, Any]],
        file_path: str
    ) -> None:
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2, default=str)

    def _compress_file(self, file_path: str) -> str:
        zip_path = f"{os.path.splitext(file_path)[0]}.zip"
        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zipf:
            zipf.write(file_path, os.path.basename(file_path))
        return zip_path

    def _process_export(
        self,
        task: ExportTask,
        compress: bool = True
    ) -> None:
        try:
            task.status = ExportStatus.PROCESSING
            self._export_tasks[task.task_id] = task

            filters = QueryFilter(**task.filters)

            if task.query_type == QueryType.LIFECYCLE:
                vuln_ids = task.filters.get("vuln_instance_ids", [])
                if not vuln_ids:
                    temp_filters = QueryFilter()
                    vuln_result = self.query_engine.query_vulnerabilities(
                        filters=temp_filters,
                        page=1,
                        page_size=10000
                    )
                    vuln_ids = [item["vuln_instance_id"] for item in vuln_result["items"]]
                data = self._get_lifecycle_data(vuln_ids, task.fields)
            else:
                data = self._get_all_data(task.query_type, filters, task.fields)

            task.total_records = len(data)

            timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
            filename = f"{task.query_type.value}_{timestamp}"

            if task.export_format == ExportFormat.CSV:
                file_path = os.path.join(config.EXPORT_TEMP_DIR, f"{filename}.csv")
                self._export_to_csv(data, file_path, task.fields)
            elif task.export_format == ExportFormat.EXCEL:
                file_path = os.path.join(config.EXPORT_TEMP_DIR, f"{filename}.xlsx")
                self._export_to_excel(data, file_path, task.query_type.value, task.fields)
            elif task.export_format == ExportFormat.JSON:
                file_path = os.path.join(config.EXPORT_TEMP_DIR, f"{filename}.json")
                self._export_to_json(data, file_path)
            else:
                raise ValueError(f"Unsupported export format: {task.export_format}")

            if compress and task.export_format != ExportFormat.JSON:
                file_path = self._compress_file(file_path)

            task.exported_records = len(data)
            task.file_path = file_path
            task.status = ExportStatus.COMPLETED
            task.completed_at = datetime.now(timezone.utc)

            self.logger.info(
                f"Export task {task.task_id} completed: {len(data)} records, file: {file_path}"
            )

        except Exception as e:
            task.status = ExportStatus.FAILED
            task.error_message = str(e)
            self.logger.error(f"Export task {task.task_id} failed: {e}")
        finally:
            self._export_tasks[task.task_id] = task

    @with_log_context(operation_type="export_data")
    def export_data(
        self,
        query_type: QueryType,
        filters: Union[QueryFilter, Dict[str, Any]],
        export_format: ExportFormat,
        fields: Optional[List[str]] = None,
        operator: str = "system",
        async_export: bool = False,
        compress: bool = True
    ) -> ExportTask:
        self.logger.info(
            f"Export request: type={query_type.value}, format={export_format.value}, operator={operator}"
        )

        if isinstance(filters, QueryFilter):
            filters_dict = asdict(filters)
        else:
            filters_dict = filters

        task_id = self._generate_task_id()
        task = ExportTask(
            task_id=task_id,
            query_type=query_type,
            export_format=export_format,
            filters=filters_dict,
            fields=fields or [],
            operator=operator
        )

        self._export_tasks[task_id] = task

        if async_export:
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            try:
                loop.run_in_executor(None, self._process_export, task, compress)
            finally:
                loop.close()
        else:
            self._process_export(task, compress)

        log_audit(
            action="export_data",
            resource_type=query_type.value,
            resource_id=task_id,
            detail=f"Format: {export_format.value}, Fields: {fields}",
            user=operator
        )

        return task

    def get_export_status(self, task_id: str) -> Optional[ExportTask]:
        task = self._export_tasks.get(task_id)
        if task:
            return task

        task_file = os.path.join(config.EXPORT_TEMP_DIR, f"{task_id}.json")
        if os.path.exists(task_file):
            try:
                with open(task_file, "r") as f:
                    task_data = json.load(f)
                return ExportTask(**task_data)
            except Exception as e:
                self.logger.error(f"Failed to read task file: {e}")

        return None

    def get_export_file_path(self, task_id: str) -> Optional[str]:
        task = self.get_export_status(task_id)
        if task and task.status == ExportStatus.COMPLETED:
            return task.file_path
        return None

    def list_export_tasks(
        self,
        operator: Optional[str] = None,
        status: Optional[ExportStatus] = None
    ) -> List[ExportTask]:
        tasks = list(self._export_tasks.values())

        if operator:
            tasks = [t for t in tasks if t.operator == operator]

        if status:
            tasks = [t for t in tasks if t.status == status]

        tasks.sort(key=lambda t: t.created_at, reverse=True)
        return tasks

    def export_lifecycle_batch(
        self,
        vuln_instance_ids: List[int],
        export_format: ExportFormat,
        fields: Optional[List[str]] = None,
        operator: str = "system",
        async_export: bool = False
    ) -> ExportTask:
        filters = {
            "vuln_instance_ids": vuln_instance_ids,
            "custom_filters": {}
        }
        return self.export_data(
            query_type=QueryType.LIFECYCLE,
            filters=filters,
            export_format=export_format,
            fields=fields,
            operator=operator,
            async_export=async_export
        )

    def batch_export_by_date_range(
        self,
        query_type: QueryType,
        start_date: datetime,
        end_date: datetime,
        export_format: ExportFormat,
        fields: Optional[List[str]] = None,
        operator: str = "system"
    ) -> List[ExportTask]:
        date_ranges = []
        current_start = start_date
        while current_start < end_date:
            current_end = min(current_start + timedelta(days=30), end_date)
            date_ranges.append((current_start, current_end))
            current_start = current_end

        tasks = []
        for i, (s, e) in enumerate(date_ranges):
            filters = QueryFilter(
                discovery_time_start=s,
                discovery_time_end=e
            )
            task = self.export_data(
                query_type=query_type,
                filters=filters,
                export_format=export_format,
                fields=fields,
                operator=operator,
                async_export=True
            )
            tasks.append(task)

        return tasks


incident_manager = IncidentManager()
incident_analyzer = IncidentAnalyzer()
timeline_manager = TimelineManager()
query_engine = QueryEngine()
export_manager = ExportManager()
